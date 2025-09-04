"""
File I/O Manager Module

Handles all Excel file reading, writing, and formatting operations.
Extracted from the main interunit_loan_matcher.py for better separation of concerns.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import os
from typing import Tuple, Dict, Any

from config import OUTPUT_FOLDER, OUTPUT_SUFFIX, SIMPLE_SUFFIX, CREATE_SIMPLE_FILES, VERBOSE_DEBUG


class FileIOManager:
    """
    Manages all file I/O operations for Excel transaction matching.
    
    Responsibilities:
    - Reading Excel files with metadata preservation
    - Writing formatted output files
    - Applying Excel styling and formatting
    - Managing output file creation
    """
    
    def __init__(self):
        """Initialize the File I/O Manager."""
        pass
    
    def read_complex_excel(self, file_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Read Excel file with metadata + transaction structure.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple of (metadata_df, transactions_df)
        """
        # Read everything first - preserve date format by reading as strings
        full_df = pd.read_excel(file_path, header=None, converters={0: str})

        # Extract metadata (rows 0-7, which are Excel rows 1-8)
        metadata = full_df.iloc[0:8, :]

        # Extract transaction data (rows 8+, which are Excel rows 9+)
        transactions = full_df.iloc[8:, :]

        # Set first row as headers and remove it from data
        transactions.columns = transactions.iloc[0]
        transactions = transactions.iloc[1:].reset_index(drop=True)

        return metadata, transactions
    
    def save_matched_files(
        self, 
        file1_matched: pd.DataFrame, 
        file2_matched: pd.DataFrame,
        metadata1: pd.DataFrame,
        metadata2: pd.DataFrame,
        file1_path: str,
        file2_path: str
    ) -> Tuple[str, str]:
        """
        Save matched files with professional formatting.
        
        Args:
            file1_matched: Matched transactions from file 1
            file2_matched: Matched transactions from file 2
            metadata1: Metadata from file 1
            metadata2: Metadata from file 2
            file1_path: Original file 1 path (for naming)
            file2_path: Original file 2 path (for naming)
            
        Returns:
            Tuple of (output_file1_path, output_file2_path)
        """
        # Generate output file paths
        base_name1 = os.path.splitext(os.path.basename(file1_path))[0]
        base_name2 = os.path.splitext(os.path.basename(file2_path))[0]
        
        output_file1 = os.path.join(OUTPUT_FOLDER, f"{base_name1}{OUTPUT_SUFFIX}")
        output_file2 = os.path.join(OUTPUT_FOLDER, f"{base_name2}{OUTPUT_SUFFIX}")
        
        # Preserve Tally date format before saving
        print("\n=== PRESERVING TALLY DATE FORMAT ===")
        self._preserve_tally_date_format(file1_matched)
        self._preserve_tally_date_format(file2_matched)
        
        # Create File 1 output
        self._create_formatted_excel_file(
            output_file1, file1_matched, metadata1
        )
        
        # Create File 2 output
        self._create_formatted_excel_file(
            output_file2, file2_matched, metadata2
        )
        
        # Create simple test files if enabled
        if CREATE_SIMPLE_FILES:
            self._create_simple_files(
                file1_matched, file2_matched, base_name1, base_name2
            )
        
        return output_file1, output_file2
    
    def _create_formatted_excel_file(
        self, 
        output_path: str, 
        matched_df: pd.DataFrame, 
        metadata_df: pd.DataFrame
    ):
        """Create a formatted Excel file with metadata and styling."""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write metadata
            metadata_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            # Write matched transactions (skip first 8 rows to make room for metadata)
            matched_df.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=8)
            
            # Get the worksheet for formatting
            worksheet = writer.sheets['Sheet1']
            
            # Apply all formatting
            self._set_column_widths(worksheet)
            self._format_amount_columns(worksheet)
            self._apply_top_alignment(worksheet)
            self._apply_filters_to_header(worksheet)
            self._apply_alternating_background_colors(worksheet, matched_df)
            self._format_output_file_transaction_blocks(worksheet)
    
    def _preserve_tally_date_format(self, transactions_df: pd.DataFrame):
        """Ensure dates are in Tally format (e.g., '01/Jul/2024') before saving."""
        if len(transactions_df.columns) > 2:  # After adding Match ID and Audit Info columns
            # Date column is now at index 2 (third column) after adding Match ID and Audit Info
            date_col = transactions_df.iloc[:, 2]  # Third column is date
            
            # Convert any datetime objects or datetime strings back to Tally format strings
            def format_tally_date(date_val):
                if pd.isna(date_val):
                    return date_val
                
                # If it's already a Tally format string, keep it
                if isinstance(date_val, str) and '/' in str(date_val) and len(str(date_val)) <= 12:
                    return date_val
                
                # If it's a datetime object, convert to Tally format
                if hasattr(date_val, 'strftime'):
                    return date_val.strftime('%d/%b/%Y')
                
                # If it's a datetime string (like '2024-07-01 00:00:00'), parse and convert
                if isinstance(date_val, str) and ('-' in str(date_val) or ':' in str(date_val)):
                    try:
                        # Parse the datetime string and convert to Tally format
                        parsed_date = pd.to_datetime(date_val)
                        return parsed_date.strftime('%d/%b/%Y')
                    except:
                        return date_val
                
                return date_val
            
            # Apply formatting to date column
            transactions_df.iloc[:, 2] = date_col.apply(format_tally_date)
            
            if VERBOSE_DEBUG:
                print(f"DEBUG: Date format preservation applied. Sample dates: {transactions_df.iloc[:3, 2].tolist()}")
    
    def _set_column_widths(self, worksheet):
        """Set professional column widths for the worksheet."""
        worksheet.column_dimensions['A'].width = 9.00
        worksheet.column_dimensions['B'].width = 30.00  # Audit Info
        worksheet.column_dimensions['C'].width = 12.00
        worksheet.column_dimensions['D'].width = 10.33
        worksheet.column_dimensions['E'].width = 60.00  # Description
        worksheet.column_dimensions['F'].width = 5.00
        worksheet.column_dimensions['G'].width = 5.00
        worksheet.column_dimensions['H'].width = 12.78
        worksheet.column_dimensions['I'].width = 9.00
        worksheet.column_dimensions['J'].width = 13.78  # Debit
        worksheet.column_dimensions['K'].width = 14.22  # Credit
        worksheet.column_dimensions['L'].width = 11.22  # Match Type
    
    def _format_amount_columns(self, worksheet):
        """Format amount columns (Debit and Credit) to prevent scientific notation."""
        debit_col = 10  # Column J (1-indexed)
        credit_col = 11  # Column K (1-indexed)
        
        # Format all data rows (starting from row 9)
        for row in range(9, worksheet.max_row + 1):
            try:
                # Format Debit column
                debit_cell = worksheet.cell(row=row, column=debit_col)
                if debit_cell.value is not None and debit_cell.value != '':
                    debit_cell.number_format = '#,##0.00'
                
                # Format Credit column
                credit_cell = worksheet.cell(row=row, column=credit_col)
                if credit_cell.value is not None and credit_cell.value != '':
                    credit_cell.number_format = '#,##0.00'
                    
            except Exception as e:
                print(f"Error formatting amount columns for row {row}: {e}")
    
    def _apply_top_alignment(self, worksheet):
        """Apply top alignment and text wrapping to ALL cells in the worksheet."""
        print(f"Setting top alignment for {worksheet.max_row} rows × {worksheet.max_column} columns...")
        
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                try:
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Always create a new alignment object to avoid style conflicts
                    new_alignment = Alignment(vertical='top')
                    
                    # Enable text wrapping for columns B (Audit Info) and E (Description)
                    if col in [2, 5]:  # Columns B and E
                        new_alignment.wrap_text = True
                    
                    # Apply the new alignment
                    cell.alignment = new_alignment
                        
                except Exception as e:
                    print(f"Error setting alignment for row {row}, col {col}: {e}")
                    continue
        
        print(f"Top alignment applied successfully!")
    
    def _apply_filters_to_header(self, worksheet):
        """Apply filters to the header row (Row 9) for easy data filtering and sorting."""
        try:
            worksheet.auto_filter.ref = f"A9:L9"
            print(f"Filters applied to header row (Row 9) successfully!")
        except Exception as e:
            print(f"Error applying filters to header row: {e}")
    
    def _apply_alternating_background_colors(self, worksheet, file_matched_df):
        """Apply alternating background colors to matched transaction blocks."""
        try:
            # Define two alternating colors
            color1 = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Light blue
            color2 = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Light yellow
            
            # Get all rows with Match IDs
            match_id_column = file_matched_df.iloc[:, 0]  # First column (Match ID)
            populated_rows = match_id_column.notna()
            
            if not populated_rows.any():
                print("No matched rows found for background coloring")
                return
            
            # Get unique Match IDs in order they appear
            unique_match_ids = []
            seen_ids = set()
            for _, match_id in enumerate(match_id_column):
                if pd.notna(match_id) and match_id not in seen_ids:
                    unique_match_ids.append(match_id)
                    seen_ids.add(match_id)
            
            print(f"Applying alternating background colors to {len(unique_match_ids)} matched transaction blocks")
            
            # Apply alternating colors to each Match ID block
            for block_index, match_id in enumerate(unique_match_ids):
                color = color1 if block_index % 2 == 0 else color2
                
                # Find all rows with this Match ID
                block_rows = file_matched_df[file_matched_df.iloc[:, 0] == match_id].index
                
                # Apply color to all rows in this block
                for df_row_idx in block_rows:
                    excel_row = df_row_idx + 10  # Convert DataFrame index to Excel row
                    
                    # Color all columns in this row
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=excel_row, column=col)
                        cell.fill = color
            
            print("Background colors applied successfully!")
            
        except Exception as e:
            print(f"Error applying background colors: {e}")
    
    def _format_output_file_transaction_blocks(self, worksheet):
        """Format output file transaction blocks: make ledger text bold, narration italic."""
        try:
            # Create fonts for different formatting
            bold_font = Font(bold=True)
            italic_font = Font(italic=True)
            bold_italic_font = Font(bold=True, italic=True)
            
            print("Formatting output file transaction blocks...")
            
            # Process all rows starting from row 10 (after metadata and header)
            for row in range(10, worksheet.max_row + 1):
                cell_d = worksheet.cell(row=row, column=4)  # Column D (Particulars)
                cell_e = worksheet.cell(row=row, column=5)  # Column E (Description)
                
                # Check if this row contains "Entered By :" in Column D
                if (cell_d.value and 
                    isinstance(cell_d.value, str) and 
                    "Entered By :" in str(cell_d.value)):
                    
                    # Make the Entered By person's name bold and italic
                    if cell_e.value:
                        cell_e.font = bold_italic_font
                
                # Check for "Opening Balance" text and make it bold
                if (cell_e.value and 
                    isinstance(cell_e.value, str) and 
                    "Opening Balance" in str(cell_e.value)):
                    
                    cell_e.font = bold_font
                    
                    # Make associated Debit and Credit values bold
                    debit_cell = worksheet.cell(row=row, column=10)  # Column J
                    credit_cell = worksheet.cell(row=row, column=11)  # Column K
                    
                    if debit_cell.value and debit_cell.value != '':
                        debit_cell.font = bold_font
                    if credit_cell.value and credit_cell.value != '':
                        credit_cell.font = bold_font
            
            print("Output file transaction block formatting completed successfully!")
            
        except Exception as e:
            print(f"Error formatting output file transaction blocks: {e}")
    
    def _create_simple_files(self, file1_matched, file2_matched, base_name1, base_name2):
        """Create simple test files without metadata."""
        simple_output1 = os.path.join(OUTPUT_FOLDER, f"{base_name1}{SIMPLE_SUFFIX}")
        simple_output2 = os.path.join(OUTPUT_FOLDER, f"{base_name2}{SIMPLE_SUFFIX}")
        
        print(f"\nCreating simple test files without metadata...")
        file1_matched.to_excel(simple_output1, index=False, header=True)
        file2_matched.to_excel(simple_output2, index=False, header=True)
        
        print(f"Created simple test files:")
        print(f"  {simple_output1}")
        print(f"  {simple_output2}")
