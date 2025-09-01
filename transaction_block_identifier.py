"""
Transaction Block Identifier Module

This module contains the logic for identifying transaction blocks in Excel files
based on specific formatting and content criteria.
"""

import openpyxl
import pandas as pd


class TransactionBlockIdentifier:
    """
    Identifies transaction blocks in Excel files based on formatting and content.
    
    Transaction Block Structure:
    1. Block Start: Date + Dr/Cr + Ledger + Vch Type (Bold) + Vch No. (Regular) + Debit/Credit (Bold)
    2. Block Content: Ledger entries (Bold) and Narration (Regular text - not bold, not italic)
    3. Block End: "Entered By :" + Person name (Bold + Italic)
    """
    
    def __init__(self):
        """Initialize the TransactionBlockIdentifier."""
        pass
    
    def get_transaction_block_rows(self, lc_match_row, file_path):
        """
        Get all row indices that belong to the transaction block containing the LC match.
        
        Args:
            lc_match_row: The row index where the LC match was found
            file_path: Path to the Excel file to analyze
        
        Returns:
            List of row indices that belong to the transaction block (from start to "Entered By :")
        """
        block_rows = []
        
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Convert DataFrame row index to Excel row number 
        # DataFrame starts at 0, but Excel has metadata rows 1-8, then headers at row 9, then data starts at row 10
        # So DataFrame row 0 = Excel row 10, DataFrame row 1 = Excel row 11, etc.
        excel_lc_row = lc_match_row + 10
        
        # FIRST: Look BACKWARDS from the LC match row to find the ACTUAL start of the transaction block
        # Transaction block starts where we find Date + Dr/Cr + Vch Type + Vch No. + Debit/Credit
        block_start_row = excel_lc_row
        for row_idx in range(excel_lc_row, 8, -1):  # Go backwards from LC row to row 9
            date_cell = ws.cell(row=row_idx, column=1)  # Column A (Date)
            particulars_cell = ws.cell(row=row_idx, column=2)  # Column B (Particulars)
            vch_type_cell = ws.cell(row=row_idx, column=6)  # Column F (Vch Type)
            vch_no_cell = ws.cell(row=row_idx, column=7)  # Column G (Vch No.)
            debit_cell = ws.cell(row=row_idx, column=8)  # Column H (Debit)
            credit_cell = ws.cell(row=row_idx, column=9)  # Column I (Credit)
            
            # Check if this row has a real date and Dr/Cr (transaction block start)
            has_real_date = (date_cell.value and 
                           str(date_cell.value).strip() and 
                           str(date_cell.value).strip() != 'None' and
                           str(date_cell.value).strip() != '')
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            
            # Check if this row has Vch Type (Bold) and Vch No. (Regular) - required for transaction blocks
            has_vch_type = vch_type_cell.value and vch_type_cell.font and vch_type_cell.font.bold
            has_vch_no = vch_no_cell.value and vch_no_cell.font and not vch_no_cell.font.bold and not vch_no_cell.font.italic
            
            # Check if this row has Debit or Credit amount (Bold) - required for transaction blocks
            has_debit = debit_cell.value and debit_cell.font and debit_cell.font.bold
            has_credit = credit_cell.value and credit_cell.font and credit_cell.font.bold
            
            # Check if this is NOT an Opening Balance row (which is not a transaction block)
            is_not_opening_balance = not (particulars_cell.value and 'Opening Balance' in str(particulars_cell.value))
            
            # Transaction block start requires: Date + Dr/Cr + Vch Type + Vch No. + Debit/Credit + NOT Opening Balance
            if (has_real_date and has_dr_cr and has_vch_type and has_vch_no and 
                (has_debit or has_credit) and is_not_opening_balance):
                # Found the start of the transaction block
                block_start_row = row_idx
                break
        
        # Convert back to DataFrame row index
        df_block_start = block_start_row - 10
        
        # SECOND: Look FORWARDS from the block start to find where it ends ("Entered By :")
        current_row = block_start_row
        while current_row <= ws.max_row:
            # Convert Excel row number to DataFrame row index
            df_row_index = current_row - 10
            if df_row_index >= 0:
                block_rows.append(df_row_index)
            
            # Check if this row contains "Entered By :" in the Particulars column (Column B)
            particulars_cell = ws.cell(row=current_row, column=2)
            if particulars_cell.value and str(particulars_cell.value).strip() == 'Entered By :':
                # Found "Entered By :", this is the end of the block
                break
            
            # Check if this row starts a NEW transaction block (Date + Dr/Cr + Vch Type + Vch No. + Debit/Credit)
            date_cell = ws.cell(row=current_row, column=1)  # Column A (Date)
            particulars_cell = ws.cell(row=current_row, column=2)  # Column B (Particulars)
            vch_type_cell = ws.cell(row=current_row, column=6)  # Column F (Vch Type)
            vch_no_cell = ws.cell(row=current_row, column=7)  # Column G (Vch No.)
            debit_cell = ws.cell(row=current_row, column=8)  # Column H (Debit)
            credit_cell = ws.cell(row=current_row, column=9)  # Column I (Credit)
            
            # Only treat as new block if it has a REAL date (not 'None' or empty) AND Dr/Cr AND Vch Type + Vch No. + Debit/Credit
            has_real_date = (date_cell.value and 
                           str(date_cell.value).strip() and 
                           str(date_cell.value).strip() != 'None' and
                           str(date_cell.value).strip() != '')
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            has_vch_type = vch_type_cell.value and vch_type_cell.font and vch_type_cell.font.bold
            has_vch_no = vch_no_cell.value and vch_no_cell.font and not vch_no_cell.font.bold and not vch_no_cell.font.italic
            has_debit = debit_cell.value and debit_cell.font and debit_cell.font.bold
            has_credit = credit_cell.value and credit_cell.font and credit_cell.font.bold
            is_not_opening_balance = not (particulars_cell.value and 'Opening Balance' in str(particulars_cell.value))
            
            # If we find a new transaction block start, stop here
            if (has_real_date and has_dr_cr and has_vch_type and has_vch_no and 
                (has_debit or has_credit) and is_not_opening_balance and current_row > block_start_row):
                # This row starts a new transaction block, so don't include it
                # The current block ends at the previous row
                break
            
            current_row += 1
        
        wb.close()
        
        print(f"DEBUG: Transaction block for LC match at row {lc_match_row} spans {len(block_rows)} rows: {block_rows}")
        print(f"DEBUG: Block starts at row {df_block_start} and includes rows up to 'Entered By :'")
        return block_rows
    
    def identify_transaction_blocks(self, transactions_df, file_path):
        """
        Identify all transaction blocks in the transactions DataFrame.
        
        Args:
            transactions_df: DataFrame containing transaction data
            file_path: Path to the Excel file to analyze formatting
        
        Returns:
            List of transaction block row indices
        """
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        transaction_blocks = []
        current_block = []
        in_block = False
        
        for row_idx in range(10, ws.max_row + 1):  # Start from row 10 (after headers)
            # Convert Excel row to DataFrame row index
            df_row_idx = row_idx - 10
            
            if df_row_idx < 0:
                continue
            
            # Check if this row starts a new transaction block
            date_cell = ws.cell(row=row_idx, column=1)  # Column A (Date)
            particulars_cell = ws.cell(row=row_idx, column=2)  # Column B (Particulars)
            vch_type_cell = ws.cell(row=row_idx, column=6)  # Column F (Vch Type)
            vch_no_cell = ws.cell(row=row_idx, column=7)  # Column G (Vch No.)
            debit_cell = ws.cell(row=row_idx, column=8)  # Column H (Debit)
            credit_cell = ws.cell(row=row_idx, column=9)  # Column I (Credit)
            
            # Check if this row has a real date and Dr/Cr
            has_real_date = (date_cell.value and 
                           str(date_cell.value).strip() and 
                           str(date_cell.value).strip() != 'None' and
                           str(date_cell.value).strip() != '')
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            
            # Check if this row has Vch Type (Bold) and Vch No. (Regular) - required for transaction blocks
            has_vch_type = vch_type_cell.value and vch_type_cell.font and vch_type_cell.font.bold
            has_vch_no = vch_no_cell.value and vch_no_cell.font and not vch_no_cell.font.bold and not vch_no_cell.font.italic
            
            # Check if this row has Debit or Credit amount (Bold) - required for transaction blocks
            has_debit = debit_cell.value and debit_cell.font and debit_cell.font.bold
            has_credit = credit_cell.value and credit_cell.font and credit_cell.font.bold
            
            # Check if this is NOT an Opening Balance row (which is not a transaction block)
            is_not_opening_balance = not (particulars_cell.value and 'Opening Balance' in str(particulars_cell.value))
            
            # Check if this row ends a transaction block ("Entered By :")
            is_block_end = particulars_cell.value and str(particulars_cell.value).strip() == 'Entered By :'
            
            # Transaction block start requires: Date + Dr/Cr + Vch Type + Vch No. + Debit/Credit + NOT Opening Balance
            is_block_start = (has_real_date and has_dr_cr and has_vch_type and has_vch_no and 
                             (has_debit or has_credit) and is_not_opening_balance)
            
            if is_block_start:
                # If we're already in a block, end the current one
                if in_block and current_block:
                    transaction_blocks.append(current_block)
                
                # Start new block
                current_block = [df_row_idx]
                in_block = True
            elif in_block:
                # Continue adding rows to current block
                current_block.append(df_row_idx)
                
                # Check if this row ends the block
                if is_block_end:
                    # End the current block
                    transaction_blocks.append(current_block)
                    current_block = []
                    in_block = False
        
        # Add the last block if it exists
        if in_block and current_block:
            transaction_blocks.append(current_block)
        
        wb.close()
        
        print(f"DEBUG: Identified {len(transaction_blocks)} transaction blocks")
        return transaction_blocks
    
    def find_transaction_block_header(self, description_row_idx, transactions_df):
        """
        Find the transaction block header row for a given description row.
        This is the UNIVERSAL method used by all matching modules.
        
        Args:
            description_row_idx: The row index of a description/narration row
            transactions_df: DataFrame containing transaction data
        
        Returns:
            Row index of the transaction block header (with date, particulars, and amounts)
        """
        # Start from the description row and go backwards to find the block header
        # Block header is the row with date and particulars (Dr/Cr) and amounts
        for row_idx in range(description_row_idx, -1, -1):
            row = transactions_df.iloc[row_idx]
            
            # Check if this row has a date and particulars
            has_date = pd.notna(row.iloc[0]) and str(row.iloc[0]).strip() != ''
            has_debit = pd.notna(row.iloc[7]) and row.iloc[7] != 0
            has_credit = pd.notna(row.iloc[8]) and row.iloc[8] != 0
            
            # Transaction block header: has date, particulars, and either debit or credit
            if has_date and (has_debit or has_credit):
                return row_idx
        
        # If no header found, return the description row itself
        return description_row_idx
    
    def find_description_row_in_block(self, row_idx, transactions_df):
        """
        Find the description row (with narration) within a transaction block.
        This is the UNIVERSAL method used by all matching modules.
        
        Args:
            row_idx: The row index to start searching from
            transactions_df: DataFrame containing transaction data
        
        Returns:
            Row index of the description row with narration text, or None if not found
        """
        # Start from the current row and look for a description row
        # Description row has narration text in column 2 (iloc[2])
        
        # Look backwards first to find description row
        for i in range(row_idx, -1, -1):
            row = transactions_df.iloc[i]
            narration = str(row.iloc[2]).strip()
            
            # Check if this row has meaningful narration text
            if (len(narration) > 10 and 
                narration.lower() not in ['nan', 'none', ''] and
                not narration.startswith('Dr') and 
                not narration.startswith('Cr')):
                return i
        
        # If not found looking backwards, look forwards
        for i in range(row_idx + 1, len(transactions_df)):
            row = transactions_df.iloc[i]
            narration = str(row.iloc[2]).strip()
            
            # Check if this row has meaningful narration text
            if (len(narration) > 10 and 
                narration.lower() not in ['nan', 'none', ''] and
                not narration.startswith('Dr') and 
                not narration.startswith('Cr')):
                return i
        
        return None
    
    def find_parent_transaction_row(self, current_row, transactions_df):
        """
        Find the parent transaction row for a description row.
        This is the UNIVERSAL method used by all matching modules.
        
        Args:
            current_row: The row index to start searching from
            transactions_df: DataFrame containing transaction data
        
        Returns:
            Row index of the parent transaction row (with date and amounts), or None if not found
        """
        # Look backwards from current row to find the most recent transaction row
        for row_idx in range(current_row - 1, -1, -1):
            row = transactions_df.iloc[row_idx]
            has_date = pd.notna(row.iloc[0])  # Date column
            has_debit = pd.notna(row.iloc[7]) and float(row.iloc[7]) > 0  # Debit column
            has_credit = pd.notna(row.iloc[8]) and float(row.iloc[8]) > 0  # Credit column
            
            if has_date and (has_debit or has_credit):
                return row_idx
        
        # If no parent found looking backwards, look forwards
        for row_idx in range(current_row + 1, len(transactions_df)):
            row = transactions_df.iloc[row_idx]
            has_date = pd.notna(row.iloc[0])  # Date column
            has_debit = pd.notna(row.iloc[7]) and float(row.iloc[7]) > 0  # Debit column
            has_credit = pd.notna(row.iloc[8]) and float(row.iloc[8]) > 0  # Credit column
            
            if has_date and (has_debit or has_credit):
                return row_idx
        
        return None
