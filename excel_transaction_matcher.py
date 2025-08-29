import pandas as pd
import numpy as np
import re
from typing import List, Dict, Any, Tuple
import logging
import os
import sys
import argparse
from openpyxl.styles import Alignment
import openpyxl
from lc_matching_logic import LCMatchingLogic

# =============================================================================
# CONFIGURATION SECTION - MODIFY THESE PATHS AS NEEDED
# =============================================================================

# Input file paths - Update these to point to your Excel files
INPUT_FILE1_PATH = "Input Files/Interunit GeoTex.xlsx"
INPUT_FILE2_PATH = "Input Files/Interunit Steel.xlsx"

# Output folder - Where to save the matched files
OUTPUT_FOLDER = "Output"

# File naming patterns for output files
OUTPUT_SUFFIX = "_MATCHED.xlsx"
SIMPLE_SUFFIX = "_SIMPLE.xlsx"

# Processing options
CREATE_SIMPLE_FILES = False  # Set to False if you don't want simple test files
CREATE_ALT_FILES = False     # Set to False if you don't want alternative files
VERBOSE_DEBUG = True         # Set to False to reduce debug output

# LC Number extraction pattern (modify if your LC numbers have different format)
LC_PATTERN = r'\b(?:L/C|LC)[-\s]?\d+[/\s]?\d*\b'

# Amount matching tolerance (for rounding differences)
AMOUNT_TOLERANCE = 0.01  # Set to 0 for exact matching, or higher for tolerance

# =============================================================================
# END CONFIGURATION SECTION
# =============================================================================

def print_configuration():
    """Print current configuration settings."""
    print("=" * 60)
    print("CURRENT CONFIGURATION")
    print("=" * 60)
    print(f"Input File 1: {INPUT_FILE1_PATH}")
    print(f"Input File 2: {INPUT_FILE2_PATH}")
    print(f"Output Folder: {OUTPUT_FOLDER}")
    print(f"Output Suffix: {OUTPUT_SUFFIX}")
    print(f"Simple Files: {'Yes' if CREATE_SIMPLE_FILES else 'No'}")
    print(f"Alternative Files: {'Yes' if CREATE_ALT_FILES else 'No'}")
    print(f"Verbose Debug: {'Yes' if VERBOSE_DEBUG else 'No'}")
    print(f"LC Pattern: {LC_PATTERN}")
    print(f"Amount Tolerance: {AMOUNT_TOLERANCE}")
    print("=" * 60)

def update_configuration():
    """Interactive configuration update (for future use)."""
    print("To update configuration, modify the variables at the top of this file:")
    print("1. INPUT_FILE1_PATH - Path to your first Excel file")
    print("2. INPUT_FILE2_PATH - Path to your second Excel file")
    print("3. OUTPUT_FOLDER - Where to save output files")
    print("4. OUTPUT_SUFFIX - Suffix for matched files")
    print("5. SIMPLE_SUFFIX - Suffix for simple test files")
    print("6. CREATE_SIMPLE_FILES - Whether to create simple test files")
    print("7. CREATE_ALT_FILES - Whether to create alternative files")
    print("8. VERBOSE_DEBUG - Whether to show detailed debug output")
    print("9. LC_PATTERN - Regex pattern for LC number extraction")
    print("10. AMOUNT_TOLERANCE - Tolerance for amount matching (0 for exact)")

class ExcelTransactionMatcher:
    """
    Handles complex Excel files with metadata rows and transaction data.
    """
    
    def __init__(self, file1_path: str, file2_path: str):
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.metadata1 = None
        self.transactions1 = None
        self.metadata2 = None
        self.transactions2 = None
        self.matching_logic = LCMatchingLogic()
        
    def read_complex_excel(self, file_path: str):
        """Read Excel file with metadata + transaction structure."""
        # Read everything first - preserve date format by reading as strings
        full_df = pd.read_excel(file_path, header=None, converters={0: str})

        # Extract metadata (rows 0-7, which are Excel rows 1-8)
        metadata = full_df.iloc[0:8, :]

        # Extract transaction data (rows 8+, which are Excel rows 9+)
        transactions = full_df.iloc[8:, :]

        # Set first row as headers and remove it from data
        transactions.columns = transactions.iloc[0]
        transactions = transactions.iloc[1:].reset_index(drop=True)

        # DEBUG: Show what columns we actually have
        print(f"DEBUG: Columns after transformation: {list(transactions.columns)}")

        # DEBUG: Show actual date values from first few rows
        print(f"DEBUG: First 5 date values (raw): {transactions.iloc[:5, 0].tolist()}")
        print(f"DEBUG: Date column data type: {transactions.iloc[0, 0].__class__.__name__}")

        return metadata, transactions
    
    def extract_lc_numbers(self, description_series):
        """Extract LC numbers from transaction descriptions."""
        def extract_single_lc(description):
            if pd.isna(description):
                return None
            
            # Pattern for LC numbers: L/C-123/456, LC-123/456, or similar formats
            match = re.search(LC_PATTERN, str(description).upper())
            return match.group() if match else None
        
        return description_series.apply(extract_single_lc)
    
    def extract_lc_numbers_with_validation(self, description_series, transactions_df):
        """Extract LC numbers and link them to parent transaction rows."""
        lc_numbers = []
        lc_parent_rows = []  # Track which parent row each LC belongs to
        
        for idx, description in enumerate(description_series):
            # Check if this row has valid transaction data
            row = transactions_df.iloc[idx]
            
            # Valid transaction row should have:
            # 1. A date (not nan)
            # 2. Either Debit or Credit amount (not both nan)
            has_date = pd.notna(row.iloc[0])  # Date column
            has_debit = pd.notna(row.iloc[7]) and float(row.iloc[7]) > 0  # Debit column
            has_credit = pd.notna(row.iloc[8]) and float(row.iloc[8]) > 0  # Credit column
            
            if has_date and (has_debit or has_credit):
                # This is a valid transaction row
                # Check if it has an LC number
                lc = self.extract_lc_numbers(pd.Series([description])).iloc[0]
                lc_numbers.append(lc)
                lc_parent_rows.append(idx)  # This row is its own parent
            else:
                # This might be a description row
                lc = self.extract_lc_numbers(pd.Series([description])).iloc[0]
                if lc is not None:
                    # Found LC in description row, need to find parent transaction row
                    parent_row = self.find_parent_transaction_row(idx, transactions_df)
                    if parent_row is not None:
                        print(f"DEBUG: LC {lc} at row {idx} linked to parent row {parent_row}")
                        lc_numbers.append(lc)
                        lc_parent_rows.append(parent_row)
                    else:
                        print(f"DEBUG: LC {lc} at row {idx} - NO PARENT FOUND!")
                        lc_numbers.append(None)
                        lc_parent_rows.append(None)
                else:
                    lc_numbers.append(None)
                    lc_parent_rows.append(None)
        
        # Store parent row mapping for later use
        self.lc_parent_mapping = dict(zip(range(len(lc_numbers)), lc_parent_rows))
        
        return pd.Series(lc_numbers)
    
    def extract_lc_numbers_from_narration(self, file_path):
        """Extract LC numbers from narration rows (non-bold Column C) using openpyxl formatting."""
        lc_numbers = []
        lc_parent_rows = []
        
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row in range(9, ws.max_row + 1):  # Start from row 9 (after headers)
            particulars_cell = ws.cell(row=row, column=2)
            desc_cell = ws.cell(row=row, column=3)
            
            # Check if this is a narration row (non-bold Column C)
            is_narration = (desc_cell.value and 
                           desc_cell.font and 
                           not desc_cell.font.bold)
            
            if is_narration:
                # This is a narration row, check for LC numbers
                narration_text = str(desc_cell.value)
                lc = self.extract_lc_numbers(pd.Series([narration_text])).iloc[0]
                
                if lc is not None:
                    # Found LC in narration row, need to find parent transaction row
                    parent_row = self.find_parent_transaction_row_with_formatting(ws, row)
                    if parent_row is not None:
                        print(f"DEBUG: LC {lc} at narration row {row} linked to parent row {parent_row}")
                        lc_numbers.append(lc)
                        lc_parent_rows.append(parent_row)
                    else:
                        print(f"DEBUG: LC {lc} at narration row {row} - NO PARENT FOUND!")
                        lc_numbers.append(None)
                        lc_parent_rows.append(None)
                else:
                    lc_numbers.append(None)
                    lc_parent_rows.append(None)
            else:
                lc_numbers.append(None)
                lc_parent_rows.append(None)
        
        wb.close()
        
        # Store parent row mapping for later use
        self.lc_parent_mapping = dict(zip(range(len(lc_numbers)), lc_parent_rows))
        
        return pd.Series(lc_numbers)
    
    def find_parent_transaction_row(self, current_row, transactions_df):
        """Find the parent transaction row for a description row."""
        # Look backwards from current row to find the most recent transaction row
        for row_idx in range(current_row - 1, -1, -1):  # Start from current_row - 1, not current_row
            row = transactions_df.iloc[row_idx]
            has_date = pd.notna(row.iloc[0])  # Date column
            has_debit = pd.notna(row.iloc[7]) and float(row.iloc[7]) > 0  # Debit column
            has_credit = pd.notna(row.iloc[8]) and float(row.iloc[8]) > 0  # Credit column
            
            if has_date and (has_debit or has_credit):
                return row_idx
        
        # If no parent found looking backwards, look forwards
        for row_idx in range(current_row + 1, len(transactions_df)):
            row = transactions_df.iloc[row_idx]
            has_date = pd.notna(row.iloc[0])  # Date column
            has_debit = pd.notna(row.iloc[7]) and float(row.iloc[7]) > 0  # Debit column
            has_credit = pd.notna(row.iloc[8]) and float(row.iloc[8]) > 0  # Credit column
            
            if has_date and (has_debit or has_credit):
                return row_idx
        
        return None
    
    def identify_transaction_blocks(self, transactions_df):
        """Identify transaction blocks based on date+Dr/Cr start and next date+Dr/Cr end."""
        blocks = []
        current_block = []
        in_block = False
        
        for idx, row in transactions_df.iterrows():
            # Check if row has date in Col A and Dr/Cr in Col B (block start/end)
            has_date = pd.notna(row.iloc[0])  # Col A (Date)
            has_dr_cr = pd.notna(row.iloc[1]) and str(row.iloc[1]).strip() in ['Dr', 'Cr']  # Col B (Particulars)
            
            # Check if this is a new block start (date + Dr/Cr)
            is_new_block_start = has_date and has_dr_cr
            
            if is_new_block_start:
                # If we're already in a block, end the current one
                if in_block and current_block:
                    blocks.append(current_block)
                
                # Start new block
                current_block = [row]
                in_block = True
            elif in_block:
                # Continue adding rows to current block
                current_block.append(row)
        
        # Add the last block if it exists
        if current_block:
            blocks.append(current_block)
        
        return blocks
    
    def identify_transaction_blocks_with_formatting(self, file_path):
        """Identify transaction blocks using openpyxl to check bold formatting in Column C."""
        blocks = []
        
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        current_block = []
        in_block = False
        
        for row in range(9, ws.max_row + 1):  # Start from row 9 (after headers)
            date_cell = ws.cell(row=row, column=1)
            particulars_cell = ws.cell(row=row, column=2)
            desc_cell = ws.cell(row=row, column=3)
            
            # Check if this is a transaction block header (Date + Dr/Cr + BOLD Col C)
            has_date = date_cell.value is not None
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            has_bold_desc = desc_cell.font and desc_cell.font.bold
            
            if has_date and has_dr_cr and has_bold_desc:
                # If we're already in a block, end the current one
                if in_block and current_block:
                    blocks.append(current_block)
                
                # Start new block
                current_block = [row]
                in_block = True
            elif in_block:
                # Continue adding rows to current block
                current_block.append(row)
        
        # Add the last block if it exists
        if current_block:
            blocks.append(current_block)
        
        wb.close()
        return blocks
    
    def process_files(self):
        """Process both files and prepare for matching."""
        print("Reading Pole Book STEEL.xlsx...")
        self.metadata1, self.transactions1 = self.read_complex_excel(self.file1_path)
        
        print("Reading Steel Book POLE.xlsx...")
        self.metadata2, self.transactions2 = self.read_complex_excel(self.file2_path)
        
        print(f"File 1: {len(self.transactions1)} rows")
        print(f"File 2: {len(self.transactions2)} rows")
        
        # DEBUG: Show column names for both files
        print(f"File 1 columns: {list(self.transactions1.columns)}")
        print(f"File 2 columns: {list(self.transactions2.columns)}")
        
        # Find the description column (should be the 3rd column, index 2)
        # Let's check what's actually in the columns
        print(f"File 1 first row: {list(self.transactions1.iloc[0, :])}")
        
        # Extract LC numbers from narration rows (non-bold Column C) using formatting
        print("Extracting LC numbers from narration rows using formatting...")
        lc_numbers1 = self.extract_lc_numbers_from_narration(self.file1_path)
        lc_numbers2 = self.extract_lc_numbers_from_narration(self.file2_path)
        
        # Identify transaction blocks using formatting
        print("Identifying transaction blocks using formatting...")
        blocks1 = self.identify_transaction_blocks_with_formatting(self.file1_path)
        blocks2 = self.identify_transaction_blocks_with_formatting(self.file2_path)
        
        print(f"File 1: {len(blocks1)} transaction blocks")
        print(f"File 2: {len(blocks2)} transaction blocks")
        
        return self.transactions1, self.transactions2, blocks1, blocks2, lc_numbers1, lc_numbers2
    
    def find_potential_matches(self):
        """Find potential LC number matches between the two files."""
        transactions1, transactions2, blocks1, blocks2, lc_numbers1, lc_numbers2 = self.process_files()
        
        # Use the new matching logic class
        matches = self.matching_logic.find_potential_matches(
            transactions1, transactions2, lc_numbers1, lc_numbers2
        )
        
        return matches
    
    def find_transaction_block_header(self, current_row, transactions_df):
        """Find the transaction block header row (with date and particulars) for a given row."""
        # Look backwards from current row to find the most recent block header
        for row_idx in range(current_row, -1, -1):
            row = transactions_df.iloc[row_idx]
            has_date = pd.notna(row.iloc[0])  # Col A (Date)
            has_particulars = pd.notna(row.iloc[1]) and str(row.iloc[1]).strip() in ['Dr', 'Cr']  # Col B (Particulars)
            
            if has_date and has_particulars:
                return row_idx
        
        # If no header found looking backwards, look forwards
        for row_idx in range(current_row + 1, len(transactions_df)):
            row = transactions_df.iloc[row_idx]
            has_date = pd.notna(row.iloc[0])  # Col A (Date)
            has_particulars = pd.notna(row.iloc[1]) and str(row.iloc[1]).strip() in ['Dr', 'Cr']  # Col B (Particulars)
            
            if has_date and has_particulars:
                return row_idx
        
        return current_row  # Fallback to current row if no header found
    
    def find_parent_transaction_row_with_formatting(self, ws, current_row):
        """Find the parent transaction row for a narration row using openpyxl formatting."""
        # Look backwards from current row to find the most recent transaction block header
        for row_idx in range(current_row, 8, -1):  # Start from current_row, go back to row 9
            date_cell = ws.cell(row=row_idx, column=1)
            particulars_cell = ws.cell(row=row_idx, column=2)
            desc_cell = ws.cell(row=row_idx, column=3)
            
            # Check if this is a transaction block header (Date + Dr/Cr + BOLD Col C)
            has_date = date_cell.value is not None
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            has_bold_desc = desc_cell.font and desc_cell.font.bold
            
            if has_date and has_dr_cr and has_bold_desc:
                return row_idx
        
        return None
    
    def get_transaction_block_rows(self, lc_match_row, file_path):
        """
        Get all row indices that belong to the transaction block containing the LC match.
        
        Args:
            lc_match_row: The row index where the LC match was found
            file_path: Path to the Excel file to analyze
        
        Returns:
            List of row indices that belong to the transaction block (from start to "Entered By :")
        """
        block_rows = []
        
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Convert DataFrame row index to Excel row number 
        # DataFrame starts at 0, but Excel has metadata rows 1-8, then headers at row 9, then data starts at row 10
        # So DataFrame row 0 = Excel row 10, DataFrame row 1 = Excel row 11, etc.
        excel_lc_row = lc_match_row + 10
        
        # FIRST: Look BACKWARDS from the LC match row to find the ACTUAL start of the transaction block
        # Transaction block starts where we find Date + Dr/Cr
        block_start_row = excel_lc_row
        for row_idx in range(excel_lc_row, 8, -1):  # Go backwards from LC row to row 9
            date_cell = ws.cell(row=row_idx, column=1)  # Column A (Date)
            particulars_cell = ws.cell(row=row_idx, column=2)  # Column B (Particulars)
            
            # Check if this row has a real date and Dr/Cr (transaction block start)
            has_real_date = (date_cell.value and 
                           str(date_cell.value).strip() and 
                           str(date_cell.value).strip() != 'None' and
                           str(date_cell.value).strip() != '')
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            
            if has_real_date and has_dr_cr:
                # Found the start of the transaction block
                block_start_row = row_idx
                break
        
        # Convert back to DataFrame row index
        df_block_start = block_start_row - 10
        
        # SECOND: Look FORWARDS from the block start to find where it ends ("Entered By :")
        current_row = block_start_row
        while current_row <= ws.max_row:
            # Convert Excel row number to DataFrame row index
            df_row_index = current_row - 10
            if df_row_index >= 0:
                block_rows.append(df_row_index)
            
            # Check if this row contains "Entered By :" in the Particulars column (Column B)
            particulars_cell = ws.cell(row=current_row, column=2)
            if particulars_cell.value and str(particulars_cell.value).strip() == 'Entered By :':
                # Found "Entered By :", this is the end of the block
                break
            
            # Check if this row starts a NEW transaction block (Date + Dr/Cr)
            date_cell = ws.cell(row=current_row, column=1)  # Column A (Date)
            particulars_cell = ws.cell(row=current_row, column=2)  # Column B (Particulars)
            
            # Only treat as new block if it has a REAL date (not 'None' or empty) AND Dr/Cr
            has_real_date = (date_cell.value and 
                           str(date_cell.value).strip() and 
                           str(date_cell.value).strip() != 'None' and
                           str(date_cell.value).strip() != '')
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            
            # If we find a new transaction block start, stop here
            if has_real_date and has_dr_cr and current_row > block_start_row:
                # This row starts a new transaction block, so don't include it
                # The current block ends at the previous row
                break
            
            current_row += 1
        
        wb.close()
        
        print(f"DEBUG: Transaction block for LC match at row {lc_match_row} spans {len(block_rows)} rows: {block_rows}")
        print(f"DEBUG: Block starts at row {df_block_start} and includes rows up to 'Entered By :'")
        return block_rows
    
    def find_narration_row_in_block(self, block_rows, file_path):
        """
        Find the first row of the narration block (where Date and DR/CR have values).
        
        Args:
            block_rows: List of row indices that belong to the transaction block
            file_path: Path to the Excel file to analyze
        
        Returns:
            Row index of the first row of the narration block, or None if not found
        """
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        narration_block_start = None
        
        # Look for the first row of the narration block (where Date and DR/CR have values)
        for block_row in block_rows:
            # Convert DataFrame row index to Excel row number (add 9 because DataFrame starts at 0, Excel starts at 9)
            excel_row = block_row + 9
            
            if excel_row <= ws.max_row:
                date_cell = ws.cell(row=excel_row, column=3)  # Column C (Date)
                debit_cell = ws.cell(row=excel_row, column=5)  # Column E (Debit)
                credit_cell = ws.cell(row=excel_row, column=6)  # Column F (Credit)
                
                # Check if this row has Date and DR/CR values (start of narration block)
                has_date = date_cell.value and str(date_cell.value).strip()
                has_debit = debit_cell.value and str(debit_cell.value).strip() and str(debit_cell.value) != 'nan'
                has_credit = credit_cell.value and str(credit_cell.value).strip() and str(credit_cell.value) != 'nan'
                
                # This is the start of narration block if it has Date and either Debit or Credit
                if has_date and (has_debit or has_credit):
                    narration_block_start = block_row
                    print(f"DEBUG: Found narration block start at Excel row {excel_row} with Date='{date_cell.value}', Debit='{debit_cell.value}', Credit='{credit_cell.value}'")
                    break
        
        wb.close()
        
        if narration_block_start is not None:
            print(f"DEBUG: Found narration block start at index {narration_block_start}")
        else:
            print(f"DEBUG: No narration block start found in transaction block")
        
        return narration_block_start

    def create_audit_info(self, match):
        """Create audit info in clean, readable plaintext format."""
        # Use the validated amounts from the matching process
        lender_amount = match['File1_Amount'] if match['File1_Type'] == 'Lender' else match['File2_Amount']
        borrower_amount = match['File1_Amount'] if match['File1_Type'] == 'Borrower' else match['File2_Amount']
        
        # Create clean, readable plaintext format
        audit_info = f"LC Match: {match['LC_Number']}\nLender Amount: {lender_amount:.2f}\nBorrower Amount: {borrower_amount:.2f}"
        return audit_info
    
    def _preserve_tally_date_format(self, transactions_df: pd.DataFrame):
        """Ensure dates are in Tally format (e.g., '01/Jul/2024') before saving."""
        if len(transactions_df.columns) > 2:  # After adding Match ID and Audit Info columns
            # Date column is now at index 2 (third column) after adding Match ID and Audit Info
            date_col = transactions_df.iloc[:, 2]  # Third column is date
            
            # Convert any datetime objects or datetime strings back to Tally format strings
            def format_tally_date(date_val):
                if pd.isna(date_val):
                    return date_val
                
                # If it's already a Tally format string, keep it
                if isinstance(date_val, str) and '/' in str(date_val) and len(str(date_val)) <= 12:
                    return date_val
                
                # If it's a datetime object, convert to Tally format
                if hasattr(date_val, 'strftime'):
                    return date_val.strftime('%d/%b/%Y')
                
                # If it's a datetime string (like '2024-07-01 00:00:00'), parse and convert
                if isinstance(date_val, str) and ('-' in str(date_val) or ':' in str(date_val)):
                    try:
                        # Parse the datetime string and convert to Tally format
                        from datetime import datetime
                        parsed_date = pd.to_datetime(date_val)
                        return parsed_date.strftime('%d/%b/%Y')
                    except:
                        return date_val
                
                return date_val
            
            # Apply formatting to date column
            transactions_df.iloc[:, 2] = date_col.apply(format_tally_date)
            
            if VERBOSE_DEBUG:
                print(f"DEBUG: Date format preservation applied. Sample dates: {transactions_df.iloc[:3, 2].tolist()}")
                print(f"DEBUG: Date column index: 2, column name: {transactions_df.columns[2]}")
                print(f"DEBUG: Date types after conversion: {[type(x) for x in transactions_df.iloc[:3, 2]]}")

    def _format_amount_columns(self, worksheet):
        """Format amount columns (Debit and Credit) to prevent scientific notation."""
        # Debit column (J) and Credit column (K) - after adding Match ID and Audit Info
        debit_col = 9  # Column J (0-indexed)
        credit_col = 10  # Column K (0-indexed)
        
        # Format all data rows (starting from row 9)
        for row in range(9, worksheet.max_row + 1):
            try:
                # Format Debit column
                debit_cell = worksheet.cell(row=row, column=debit_col + 1)  # openpyxl uses 1-indexed
                if debit_cell.value is not None and debit_cell.value != '':
                    debit_cell.number_format = '#,##0.00'
                
                # Format Credit column
                credit_cell = worksheet.cell(row=row, column=credit_col + 1)  # openpyxl uses 1-indexed
                if credit_cell.value is not None and credit_cell.value != '':
                    credit_cell.number_format = '#,##0.00'
                    
            except Exception as e:
                print(f"Error formatting amount columns for row {row}: {e}")

    def _set_column_widths(self, worksheet):
        """Set column widths for the worksheet"""
        worksheet.column_dimensions['A'].width = 9.00
        worksheet.column_dimensions['B'].width = 30.00
        worksheet.column_dimensions['C'].width = 12.00
        worksheet.column_dimensions['D'].width = 10.33
        worksheet.column_dimensions['E'].width = 60.00
        worksheet.column_dimensions['F'].width = 5.00
        worksheet.column_dimensions['G'].width = 5.00
        worksheet.column_dimensions['H'].width = 12.78
        worksheet.column_dimensions['I'].width = 9.00
        worksheet.column_dimensions['J'].width = 12.78
        worksheet.column_dimensions['K'].width = 14.22

    def create_matched_files(self, matches, transactions1, transactions2):
        """Create matched versions of both files with new columns."""
        if not matches:
            print("No matches found. Cannot create matched files.")
            return
        
        # Create file1 with new columns
        file1_matched = transactions1.copy()
        
        # Insert new columns at the beginning
        file1_matched.insert(0, 'Match ID', None)
        file1_matched.insert(1, 'Audit Info', None)
        

        
        print(f"DEBUG: File1 DataFrame created with shape: {file1_matched.shape}")
        print(f"DEBUG: File1 columns: {list(file1_matched.columns)}")
        
        # Create file2 with new columns
        file2_matched = transactions2.copy()
        
        # Insert new columns at the beginning
        file2_matched.insert(0, 'Match ID', None)
        file2_matched.insert(1, 'Audit Info', None)
        
        print(f"DEBUG: File2 DataFrame created with shape: {file2_matched.shape}")
        print(f"DEBUG: File2 columns: {list(file2_matched.columns)}")
        
        # Verify the new columns are actually there
        print(f"DEBUG: File1 first few rows of Match ID column:")
        print(file1_matched.iloc[:5, 0].tolist())
        print(f"DEBUG: File1 first few rows of Audit Info column:")
        print(file1_matched.iloc[:5, 1].tolist())
        
        print(f"\n=== DEBUG: MATCH DATA POPULATION ===")
        
        # Populate match information
        for match in matches:
            match_id = match['match_id']  # Use the pre-assigned match ID
            audit_info = self.create_audit_info(match)
            
            print(f"Match {match_id}:")
            print(f"  LC Number: {match['LC_Number']}")
            print(f"  File1 Row {match['File1_Index']}: Debit={match['File1_Debit']}, Credit={match['File1_Credit']}")
            print(f"  File2 Row {match['File2_Index']}: Debit={match['File2_Debit']}, Credit={match['File2_Credit']}")
            print(f"  Audit Info: {audit_info}")
            
            # Update file1 - populate entire transaction block with Match ID and Audit Info
            file1_row_idx = match['File1_Index']
            print(f"    DEBUG: Setting File1 row {file1_row_idx} col 0 to '{match_id}'")
            print(f"    DEBUG: Setting File1 row {file1_row_idx} col 1 to '{audit_info[:50]}...'")
            
            # Find the entire transaction block for file1 and populate all rows
            file1_block_rows = self.get_transaction_block_rows(file1_row_idx, self.file1_path)
            print(f"    DEBUG: File1 transaction block spans rows: {file1_block_rows}")
            
            # Populate ALL rows of the transaction block with Match ID, but Audit Info only in second-to-last row
            for i, block_row in enumerate(file1_block_rows):
                if 0 <= block_row < len(file1_matched):
                    file1_matched.iloc[block_row, 0] = match_id  # Match ID column on all rows
                    
                    # Audit Info goes ONLY in the second-to-last row of the transaction block
                    if i == len(file1_block_rows) - 2:  # Second-to-last row
                        file1_matched.iloc[block_row, 1] = audit_info
                        print(f"    DEBUG: Populated File1 row {block_row} with Match ID '{match_id}' and Audit Info (second-to-last row)")
                    else:
                        print(f"    DEBUG: Populated File1 row {block_row} with Match ID '{match_id}' only")
            

            
            # Update file2 - populate entire transaction block with Match ID and Audit Info
            file2_row_idx = match['File2_Index']
            print(f"    DEBUG: Setting File2 row {file2_row_idx} col 0 to '{match_id}'")
            print(f"    DEBUG: Setting File2 row {file2_row_idx} col 1 to '{audit_info[:50]}...'")
            
            # Find the entire transaction block for file2 and populate all rows
            file2_block_rows = self.get_transaction_block_rows(file2_row_idx, self.file2_path)
            print(f"    DEBUG: File2 transaction block spans rows: {file2_block_rows}")
            
            # Populate ALL rows of the transaction block with Match ID, but Audit Info only in second-to-last row
            for i, block_row in enumerate(file2_block_rows):
                if 0 <= block_row < len(file2_matched):
                    file2_matched.iloc[block_row, 0] = match_id  # Match ID column on all rows
                    
                    # Audit Info goes ONLY in the second-to-last row of the transaction block
                    if i == len(file2_block_rows) - 2:  # Second-to-last row
                        file2_matched.iloc[block_row, 1] = audit_info
                        print(f"    DEBUG: Populated File2 row {block_row} with Match ID '{match_id}' and Audit Info (second-to-last row)")
                    else:
                        print(f"    DEBUG: Populated File2 row {block_row} with Match ID '{match_id}' only")
        
        # Save matched files using configuration variables
        base_name1 = os.path.splitext(os.path.basename(self.file1_path))[0]
        base_name2 = os.path.splitext(os.path.basename(self.file2_path))[0]
        
        output_file1 = os.path.join(OUTPUT_FOLDER, f"{base_name1}{OUTPUT_SUFFIX}")
        output_file2 = os.path.join(OUTPUT_FOLDER, f"{base_name2}{OUTPUT_SUFFIX}")
        
        if VERBOSE_DEBUG:
            print(f"\n=== DEBUG: BEFORE SAVING ===")
            print(f"File1 - Rows with Match IDs: {file1_matched.iloc[:, 0].notna().sum()}")
            print(f"File1 - Rows with Audit Info: {file1_matched.iloc[:, 1].notna().sum()}")
            print(f"File2 - Rows with Match IDs: {file2_matched.iloc[:, 0].notna().sum()}")
            print(f"File2 - Rows with Audit Info: {file2_matched.iloc[:, 1].notna().sum()}")
            
            # Show some actual values to verify they're there
            print(f"\n=== DEBUG: ACTUAL VALUES IN DATAFRAME ===")
            
            # Get the actual populated rows dynamically
            populated_rows = file1_matched.iloc[:, 0].notna()
            if populated_rows.any():
                populated_indices = file1_matched[populated_rows].index
                for idx in populated_indices[:4]:  # Show first 4 populated rows
                    print(f"File1 - Row {idx} Match ID: '{file1_matched.iloc[idx, 0]}'")
                    print(f"File1 - Row {idx} Audit Info: '{file1_matched.iloc[idx, 1]}'")
            else:
                print("No populated rows found in File1")
        
        # Preserve Tally date format before saving
        print("\n=== PRESERVING TALLY DATE FORMAT ===")
        self._preserve_tally_date_format(file1_matched)
        self._preserve_tally_date_format(file2_matched)
        
        # Create output with metadata + matched transactions
        with pd.ExcelWriter(output_file1, engine='openpyxl') as writer:
            # Write metadata
            self.metadata1.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            # Write matched transactions (skip first 8 rows to make room for metadata)
            file1_matched.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=8)
            
            # Get the worksheet to set column widths
            worksheet = writer.sheets['Sheet1']
            self._set_column_widths(worksheet)
            self._format_amount_columns(worksheet) # Apply amount formatting
            
            # Enable text wrapping for columns B (Audit Info) and E (Description)
            # Note: wrap_text is a cell property, not column property
            # We'll set it on ALL rows that contain data
            for row in range(9, worksheet.max_row + 1):  # ALL rows from 9 to max
                try:
                    # Column B (Audit Info)
                    cell_b = worksheet.cell(row=row, column=2)
                    if cell_b.value:
                        cell_b.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Column E (Description)
                    cell_e = worksheet.cell(row=row, column=5)
                    if cell_e.value:
                        cell_e.alignment = Alignment(wrap_text=True, vertical='top')
                        
                except Exception as e:
                    print(f"Error setting text wrapping for row {row}: {e}")
            
                    
        with pd.ExcelWriter(output_file2, engine='openpyxl') as writer:
            # Write metadata
            self.metadata2.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            # Write matched transactions (skip first 8 rows to make room for metadata)
            file2_matched.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=8)
            
            # Get the worksheet to set column widths
            worksheet = writer.sheets['Sheet1']
            self._set_column_widths(worksheet)
            self._format_amount_columns(worksheet) # Apply amount formatting
            
            # Enable text wrapping for columns B (Audit Info) and E (Description)
            # Note: wrap_text is a cell property, not column property
            # We'll set it on ALL rows that contain data
            for row in range(9, worksheet.max_row + 1):  # ALL rows from 9 to max
                try:
                    # Column B (Audit Info)
                    cell_b = worksheet.cell(row=row, column=2)
                    if cell_b.value:
                        cell_b.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Column E (Description)
                    cell_e = worksheet.cell(row=row, column=5)
                    if cell_e.value:
                        cell_e.alignment = Alignment(wrap_text=True, vertical='top')
                        
                except Exception as e:
                    print(f"Error setting text wrapping for row {row}: {e}")
            
        
        # Also create a simple version without metadata to test (if enabled)
        if CREATE_SIMPLE_FILES:
            simple_output1 = os.path.join(OUTPUT_FOLDER, f"{base_name1}{SIMPLE_SUFFIX}")
            simple_output2 = os.path.join(OUTPUT_FOLDER, f"{base_name2}{SIMPLE_SUFFIX}")
            
            print(f"\nCreating simple test files without metadata...")
            file1_matched.to_excel(simple_output1, index=False, header=True)
            file2_matched.to_excel(simple_output2, index=False, header=True)
            
            print(f"Created simple test files:")
            print(f"  {simple_output1}")
            print(f"  {simple_output2}")
        

        
        print(f"\n=== DEBUG: AFTER SAVING ===")
        print(f"Checking if files were actually written...")
        
        # Verify the files were written correctly
        try:
            df_check1 = pd.read_excel(output_file1, header=8)
            print(f"File1 loaded successfully, shape: {df_check1.shape}")
            print(f"File1 - Rows with Match IDs: {df_check1.iloc[:, 0].notna().sum()}")
            print(f"File1 - Rows with Audit Info: {df_check1.iloc[:, 1].notna().sum()}")
            
            # Check if text wrapping was applied by reading the Excel file with openpyxl
            print(f"\n=== VERIFYING TEXT WRAPPING IN FILE 1 ===")
            wb1 = openpyxl.load_workbook(output_file1)
            ws1 = wb1.active
            print(f"Worksheet: {ws1.title}")
            print(f"Max row: {ws1.max_row}, Max column: {ws1.max_column}")
            
            # Check a few cells in columns B and E for text wrapping
            for row in range(9, min(15, ws1.max_row + 1)):
                cell_b = ws1.cell(row=row, column=2)
                cell_e = ws1.cell(row=row, column=5)
                print(f"Row {row}:")
                print(f"  Column B: value='{cell_b.value}', wrap_text={cell_b.alignment.wrap_text if cell_b.alignment else 'None'}")
                print(f"  Column E: value='{cell_e.value}', wrap_text={cell_e.alignment.wrap_text if cell_e.alignment else 'None'}")
            
            wb1.close()
            
        except Exception as e:
            print(f"Error reading File1: {e}")
        
        try:
            df_check2 = pd.read_excel(output_file2, header=8)
            print(f"File2 loaded successfully, shape: {df_check2.shape}")
            print(f"File2 - Rows with Match IDs: {df_check2.iloc[:, 0].notna().sum()}")
            print(f"File2 - Rows with Audit Info: {df_check2.iloc[:, 1].notna().sum()}")
            
            # Check if text wrapping was applied by reading the Excel file with openpyxl
            print(f"\n=== VERIFYING TEXT WRAPPING IN FILE 2 ===")
            wb2 = openpyxl.load_workbook(output_file2)
            ws2 = wb2.active
            print(f"Worksheet: {ws2.title}")
            print(f"Max row: {ws2.max_row}, Max column: {ws2.max_column}")
            
            # Check a few cells in columns B and E for text wrapping
            for row in range(9, min(15, ws2.max_row + 1)):
                cell_b = ws2.cell(row=row, column=2)
                cell_e = ws2.cell(row=row, column=5)
                print(f"Row {row}:")
                print(f"  Column B: value='{cell_b.value}', wrap_text={cell_b.alignment.wrap_text if cell_b.alignment else 'None'}")
                print(f"  Column E: value='{cell_e.value}', wrap_text={cell_e.alignment.wrap_text if cell_e.alignment else 'None'}")
            
            wb2.close()
            
        except Exception as e:
            print(f"Error reading File2: {e}")
        
        print(f"\nCreated matched files:")
        print(f"  {output_file1}")
        print(f"  {output_file2}")
        
        # Verify the data was populated correctly
        self.verify_match_data(file1_matched, file2_matched, matches)
    

    
    def verify_match_data(self, file1_matched, file2_matched, matches):
        """Verify that Match ID and Audit Info columns are properly populated."""
        print(f"\n=== VERIFICATION RESULTS ===")
        
        # Check Match ID column population - only count non-empty, non-NaN values
        match_ids_1 = file1_matched.iloc[:, 0].replace('', None).dropna()
        match_ids_2 = file2_matched.iloc[:, 0].replace('', None).dropna()
        
        print(f"File 1 - Match IDs populated: {len(match_ids_1)}")
        print(f"File 2 - Match IDs populated: {len(match_ids_2)}")
        
        # Check Audit Info column population - only count non-empty, non-NaN values
        audit_info_1 = file1_matched.iloc[:, 1].replace('', None).dropna()
        audit_info_2 = file2_matched.iloc[:, 1].replace('', None).dropna()
        
        print(f"File 1 - Audit Info populated: {len(audit_info_1)}")
        print(f"File 2 - Audit Info populated: {len(audit_info_2)}")
        
        # Show sample populated data
        if len(match_ids_1) > 0:
            print(f"\nFile 1 - Sample populated rows:")
            for i, match_id in enumerate(match_ids_1[:3]):
                # Find rows that actually have this match_id (not empty strings)
                mask = (file1_matched.iloc[:, 0] == match_id) & (file1_matched.iloc[:, 0] != '')
                if mask.any():
                    row_idx = file1_matched[mask].index[0]
                    print(f"  Row {row_idx}: Match ID = {match_id}")
                    print(f"    Date: {file1_matched.iloc[row_idx, 2]}")
                    print(f"    Description: {str(file1_matched.iloc[row_idx, 3])[:50]}...")
                    print(f"    Debit: {file1_matched.iloc[row_idx, 9]}, Credit: {file1_matched.iloc[row_idx, 10]}")
        
        if len(match_ids_2) > 0:
            print(f"\nFile 2 - Sample populated rows:")
            for i, match_id in enumerate(match_ids_2[:3]):
                # Find rows that actually have this match_id (not empty strings)
                mask = (file2_matched.iloc[:, 0] == match_id) & (file2_matched.iloc[:, 0] != '')
                if mask.any():
                    row_idx = file2_matched[mask].index[0]
                    print(f"  Row {row_idx}: Match ID = {match_id}")
                    print(f"    Date: {file2_matched.iloc[row_idx, 2]}")
                    print(f"    Description: {str(file2_matched.iloc[row_idx, 3])[:50]}...")
                    print(f"    Debit: {file2_matched.iloc[row_idx, 9]}, Credit: {file2_matched.iloc[row_idx, 10]}")

def main():
    # Show current configuration
    print_configuration()
    print()
    
    # Use configuration variables from the top of the file
    print(f"=== PROCESSING FILES ===")
    
    # Check if input files exist
    if not os.path.exists(INPUT_FILE1_PATH):
        print(f"ERROR: Input file 1 not found: {INPUT_FILE1_PATH}")
        return
    if not os.path.exists(INPUT_FILE2_PATH):
        print(f"ERROR: Input file 2 not found: {INPUT_FILE2_PATH}")
        return
    
    # Create output folder if it doesn't exist
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    # Create matcher instance
    matcher = ExcelTransactionMatcher(INPUT_FILE1_PATH, INPUT_FILE2_PATH)
    matches = matcher.find_potential_matches()
    transactions1, transactions2, blocks1, blocks2, lc_numbers1, lc_numbers2 = matcher.process_files()
    
    print(f"\n=== SUMMARY ===")
    print(f"Total potential matches found: {len(matches)}")
    
    if matches:
        print("\nCreating matched output files...")
        matcher.create_matched_files(matches, transactions1, transactions2)
        print("\nOutput files created successfully!")
    else:
        print("\nNo matches found. No output files created.")

if __name__ == "__main__":
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Excel Transaction Matcher for LC Numbers')
    parser.add_argument('--file1', help='Path to first Excel file (overrides config)')
    parser.add_argument('--file2', help='Path to second Excel file (overrides config)')
    parser.add_argument('--output', help='Output folder (overrides config)')
    parser.add_argument('--config', action='store_true', help='Show current configuration')
    parser.add_argument('--help-config', action='store_true', help='Show configuration help')
    
    args = parser.parse_args()
    
    # Handle special arguments
    if args.help_config:
        update_configuration()
        sys.exit(0)
    
    if args.config:
        print_configuration()
        sys.exit(0)
    
    # Override configuration with command line arguments if provided
    if args.file1:
        INPUT_FILE1_PATH = args.file1
    if args.file2:
        INPUT_FILE2_PATH = args.file2
    if args.output:
        OUTPUT_FOLDER = args.output
    
    # Run the main function
    main()
    