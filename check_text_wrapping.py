import openpyxl

def check_text_wrapping(file_path):
    print(f"\n=== CHECKING TEXT WRAPPING IN: {file_path} ===")
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        print(f"Worksheet: {ws.title}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        # Check a few cells in columns B and E for text wrapping
        for row in range(9, min(15, ws.max_row + 1)):
            cell_b = ws.cell(row=row, column=2)
            cell_e = ws.cell(row=row, column=5)
            print(f"Row {row}:")
            print(f"  Column B: value='{cell_b.value}', wrap_text={cell_b.alignment.wrap_text if cell_b.alignment else 'None'}")
            print(f"  Column E: value='{cell_e.value}', wrap_text={cell_e.alignment.wrap_text if cell_e.alignment else 'None'}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error reading file: {e}")

# Check both files
check_text_wrapping("Output/Interunit Steel_MATCHED.xlsx")
check_text_wrapping("Output/Interunit GeoTex_MATCHED.xlsx")
