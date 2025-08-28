import pandas as pd
import openpyxl

# Check the actual Excel file structure around the rows we're working with
wb = openpyxl.load_workbook("Input Files/Interunit Steel.xlsx")
ws = wb.active

print("=== VERIFYING ROW MAPPING ===")
print("DataFrame row 21 should map to Excel row 31 (based on our calculation)")

# Check Excel rows around 31
print("\n=== EXCEL ROWS AROUND 31 (DataFrame row 21) ===")
for row_num in range(29, 35):
    date_val = ws.cell(row=row_num, column=1).value
    particulars_val = ws.cell(row=row_num, column=2).value
    desc_val = ws.cell(row=row_num, column=3).value
    
    print(f"Excel Row {row_num}: Date='{date_val}', Particulars='{particulars_val}', Desc='{str(desc_val)[:50]}...'")
    
    # Check if this is "Entered By :"
    if particulars_val and str(particulars_val).strip() == 'Entered By :':
        print(f"  *** FOUND 'Entered By :' at Excel row {row_num} ***")
    
    # Check if this looks like a transaction start
    has_date = date_val and str(date_val).strip() and str(date_val).strip() != 'None' and str(date_val).strip() != ''
    has_dr_cr = particulars_val and str(particulars_val).strip() in ['Dr', 'Cr']
    if has_date and has_dr_cr:
        print(f"  *** LOOKS LIKE TRANSACTION START at Excel row {row_num} ***")

# Check what should be the transaction block for row 31
print("\n=== EXPECTED TRANSACTION BLOCK FOR ROW 31 ===")
print("Row 31: GULC#308524022796/24-Nut & Bolt (should be M001)")
print("Should include rows 31, 32, 33 (up to 'Entered By :')")

wb.close()
