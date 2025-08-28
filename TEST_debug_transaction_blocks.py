import pandas as pd
import openpyxl
import sys
import os

# Add current directory to path to import the class
sys.path.append('.')

from excel_transaction_matcher import ExcelTransactionMatcher

# Create matcher instance
matcher = ExcelTransactionMatcher("Input Files/Interunit Steel.xlsx", "Input Files/Interunit GeoTex.xlsx")

# Test with different rows to find a multi-row transaction block
test_rows = [21, 28, 31, 34, 40]
print("Testing transaction block expansion for different rows:")

for test_row in test_rows:
    print(f"\n--- Testing row {test_row} ---")
    
    # Test the method directly
    block_rows = matcher.get_transaction_block_rows(test_row, "Input Files/Interunit Steel.xlsx")
    print(f"Result: {block_rows}")
    
    if len(block_rows) > 1:
        print(f"*** FOUND MULTI-ROW BLOCK! Spans {len(block_rows)} rows ***")
        break

# Let's also check what's in the Excel file around row 28 (which had a multi-row structure)
wb = openpyxl.load_workbook("Input Files/Interunit Steel.xlsx")
ws = wb.active

test_row = 28
print(f"\n=== EXCEL FILE CONTENT AROUND ROW {test_row + 9} ===")
print("Looking at rows to understand transaction structure:")
for row_num in range(test_row + 7, test_row + 15):  # Check rows around 28
    if row_num <= ws.max_row:
        date_val = ws.cell(row=row_num, column=1).value
        particulars_val = ws.cell(row=row_num, column=2).value
        desc_val = ws.cell(row=row_num, column=3).value
        
        # Show the actual values
        date_str = str(date_val) if date_val is not None else 'None'
        particulars_str = str(particulars_val) if particulars_val is not None else 'None'
        desc_str = str(desc_val)[:50] + '...' if desc_val and len(str(desc_val)) > 50 else str(desc_val)
        
        print(f"Row {row_num}: Date='{date_str}', Particulars='{particulars_str}', Desc='{desc_str}'")
        
        # Check if this is "Entered By :"
        if particulars_val and str(particulars_val).strip() == 'Entered By :':
            print(f"  *** FOUND 'Entered By :' at row {row_num} ***")

wb.close()
