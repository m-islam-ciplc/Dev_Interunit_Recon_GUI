import pandas as pd
import openpyxl

# Check the actual Excel file structure
wb = openpyxl.load_workbook("Input Files/Interunit Steel.xlsx")
ws = wb.active

print("=== EXCEL FILE STRUCTURE ===")
print(f"Total rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")

# Check the first few rows to understand the structure
print("\n=== FIRST 15 ROWS ===")
for row_num in range(1, 16):
    date_val = ws.cell(row=row_num, column=1).value
    particulars_val = ws.cell(row=row_num, column=2).value
    desc_val = ws.cell(row=row_num, column=3).value
    
    print(f"Excel Row {row_num}: Date='{date_val}', Particulars='{particulars_val}', Desc='{desc_val}'")
    
    # Check if this is "Entered By :"
    if particulars_val and str(particulars_val).strip() == 'Entered By :':
        print(f"  *** FOUND 'Entered By :' at Excel row {row_num} ***")

# Check around row 30 (which should be the first transaction block)
print("\n=== AROUND ROW 30 (FIRST TRANSACTION BLOCK) ===")
for row_num in range(25, 35):
    date_val = ws.cell(row=row_num, column=1).value
    particulars_val = ws.cell(row=row_num, column=2).value
    desc_val = ws.cell(row=row_num, column=3).value
    
    print(f"Excel Row {row_num}: Date='{date_val}', Particulars='{particulars_val}', Desc='{desc_val}'")
    
    # Check if this is "Entered By :"
    if particulars_val and str(particulars_val).strip() == 'Entered By :':
        print(f"  *** FOUND 'Entered By :' at Excel row {row_num} ***")
    
    # Check if this looks like a transaction start
    has_date = date_val and str(date_val).strip() and str(date_val).strip() != 'None' and str(date_val).strip() != ''
    has_dr_cr = particulars_val and str(particulars_val).strip() in ['Dr', 'Cr']
    if has_date and has_dr_cr:
        print(f"  *** LOOKS LIKE TRANSACTION START at Excel row {row_num} ***")

wb.close()
