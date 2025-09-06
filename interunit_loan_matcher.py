import pandas as pd
import re

import os
import sys
import argparse
from openpyxl.styles import Alignment
import openpyxl
from openpyxl import load_workbook
from matching_logic import (
    LCMatchingLogic, POMatchingLogic, USDMatchingLogic,
    InterunitLoanMatcher, AggregatedPOMatchingLogic, NarrationMatchingLogic
)
from transaction_block_identifier import TransactionBlockIdentifier

# =============================================================================
# CONFIGURATION SECTION
# =============================================================================
# Import configuration from dedicated config module
from config import (
    INPUT_FILE1_PATH, INPUT_FILE2_PATH, OUTPUT_FOLDER, OUTPUT_SUFFIX,
    SIMPLE_SUFFIX, CREATE_SIMPLE_FILES, 
    VERBOSE_DEBUG
)

# Import patterns from the matching logic package
from matching_logic import LC_PATTERN, PO_PATTERN, USD_PATTERN

def print_configuration():
    """Print current configuration settings."""
    print("=" * 60)
    print("CURRENT CONFIGURATION")
    print("=" * 60)
    print(f"Input File 1: {INPUT_FILE1_PATH}")
    print(f"Input File 2: {INPUT_FILE2_PATH}")
    print(f"Output Folder: {OUTPUT_FOLDER}")
    print(f"Output Suffix: {OUTPUT_SUFFIX}")
    print(f"Simple Files: {'Yes' if CREATE_SIMPLE_FILES else 'No'}")

    print(f"Verbose Debug: {'Yes' if VERBOSE_DEBUG else 'No'}")
    print(f"LC Pattern: {LC_PATTERN}")
    print(f"PO Pattern: {PO_PATTERN}")
    print(f"USD Pattern: {USD_PATTERN}")
    print("Narration Matching: Enabled (Exact text matching)")

    print("=" * 60)

def update_configuration():
    """Interactive configuration update (for future use)."""
    print("To update configuration, modify the variables at the top of this file:")
    print("1. INPUT_FILE1_PATH - Path to your first Excel file")
    print("2. INPUT_FILE2_PATH - Path to your second Excel file")
    print("3. OUTPUT_FOLDER - Where to save output files")
    print("4. OUTPUT_SUFFIX - Suffix for matched files")
    print("5. SIMPLE_SUFFIX - Suffix for simple test files")
    print("6. CREATE_SIMPLE_FILES - Whether to create simple test files")

    print("8. VERBOSE_DEBUG - Whether to show detailed debug output")
    print("9. LC_PATTERN - Regex pattern for LC number extraction (defined in lc_matching_logic.py)")
    print("10. PO_PATTERN - Regex pattern for PO number extraction (defined in po_matching_logic.py)")
    print("11. USD_PATTERN - Regex pattern for USD amount extraction (defined in usd_matching_logic.py)")
    print("12. Narration Matching - Exact text matching between files (highest priority)")


class   ExcelTransactionMatcher:
    """
    Handles complex Excel files with metadata rows and transaction data.
    """
    
    def __init__(self, file1_path: str, file2_path: str):
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.metadata1 = None
        self.transactions1 = None
        self.metadata2 = None
        self.transactions2 = None
        self.block_identifier = TransactionBlockIdentifier()
        self.lc_matching_logic = LCMatchingLogic(self.block_identifier)
        self.po_matching_logic = POMatchingLogic(self.block_identifier)
        self.usd_matching_logic = USDMatchingLogic(self.block_identifier)
        self.interunit_loan_matcher = InterunitLoanMatcher(self.block_identifier)
        self.aggregated_po_matching_logic = AggregatedPOMatchingLogic(self.block_identifier)
        self.narration_matching_logic = NarrationMatchingLogic(self.block_identifier)
        
        # Performance optimization caches
        self._block_header_cache1 = {}
        self._block_header_cache2 = {}
        self._amount_cache1 = {}
        self._amount_cache2 = {}
        
        # Cached workbook data for performance
        self._cached_wb1 = None
        self._cached_ws1 = None
        self._cached_wb2 = None
        self._cached_ws2 = None
        self._cached_blocks1 = None
        self._cached_blocks2 = None
        self._cached_formatting_data1 = None
        self._cached_formatting_data2 = None
        
        # Compiled regex patterns for performance
        self._compiled_lc_pattern = re.compile(LC_PATTERN)
        self._compiled_po_pattern = re.compile(PO_PATTERN)
        self._compiled_usd_pattern = re.compile(USD_PATTERN)
        self._compiled_interunit_pattern = re.compile(r'([A-Z]{2,4})#(\d{4,6})')
        
        # Cached extracted data
        self._cached_extracted_data = None
        
        # Optimized data access caches
        self._lc_numbers1_array = None
        self._lc_numbers2_array = None
        self._po_numbers1_array = None
        self._po_numbers2_array = None
        self._usd_amounts1_array = None
        self._usd_amounts2_array = None
        self._interunit_accounts1_array = None
        self._interunit_accounts2_array = None
        
        # Additional performance caches
        self._block_header_cache = {}  # Universal block header cache
        self._description_row_cache = {}  # Universal description row cache
        self._narration_cache = {}  # Cached narration strings
        self._amount_cache = {}  # Universal amount cache
        
    def read_complex_excel(self, file_path: str):
        """Read Excel file with metadata + transaction structure."""
        # Read everything as strings to preserve all formatting
        full_df = pd.read_excel(file_path, header=None, dtype=str)

        # Extract metadata (rows 0-7, which are Excel rows 1-8)
        metadata = full_df.iloc[0:8, :]

        # Extract transaction data (rows 8+, which are Excel rows 9+)
        transactions = full_df.iloc[8:, :]

        # Set first row as headers and remove it from data
        transactions.columns = transactions.iloc[0]
        transactions = transactions.iloc[1:].reset_index(drop=True)

        return metadata, transactions
    
    def extract_amounts_from_strings(self, row):
        """Extract amounts from row data that's loaded as strings - OPTIMIZED."""
        # Get raw values once
        debit_raw = row.iloc[7] if pd.notna(row.iloc[7]) else '0'
        credit_raw = row.iloc[8] if pd.notna(row.iloc[8]) else '0'
        
        # Convert to strings once
        debit_str = str(debit_raw)
        credit_str = str(credit_raw)
        
        # Optimized numeric conversion
        try:
            # Remove commas and check if numeric in one pass
            debit_clean = debit_str.replace(',', '')
            credit_clean = credit_str.replace(',', '')
            
            # Use faster numeric check
            debit = float(debit_clean) if debit_clean.replace('.', '').isdigit() else 0.0
            credit = float(credit_clean) if credit_clean.replace('.', '').isdigit() else 0.0
        except (ValueError, TypeError):
            debit, credit = 0.0, 0.0
        
        return debit, credit
    
    def load_and_cache_workbooks(self):
        """Load workbooks once and cache them for performance."""
        if self._cached_wb1 is None:
            print("Loading and caching workbooks for performance...")
            self._cached_wb1 = load_workbook(self.file1_path)
            self._cached_ws1 = self._cached_wb1.active
            self._cached_wb2 = load_workbook(self.file2_path)
            self._cached_ws2 = self._cached_wb2.active
            print("Workbooks cached successfully")
    
    def get_cached_blocks(self, file_num):
        """Get cached transaction blocks or identify them if not cached."""
        if file_num == 1:
            if self._cached_blocks1 is None:
                print("Identifying transaction blocks for File 1...")
                self._cached_blocks1 = self.block_identifier.identify_transaction_blocks(
                    self.transactions1, self.file1_path
                )
                print(f"File 1: {len(self._cached_blocks1)} transaction blocks cached")
            return self._cached_blocks1
        else:
            if self._cached_blocks2 is None:
                print("Identifying transaction blocks for File 2...")
                self._cached_blocks2 = self.block_identifier.identify_transaction_blocks(
                    self.transactions2, self.file2_path
                )
                print(f"File 2: {len(self._cached_blocks2)} transaction blocks cached")
            return self._cached_blocks2
    
    def get_cached_formatting_data(self, file_num):
        """Get cached formatting data or analyze it if not cached."""
        if file_num == 1:
            if self._cached_formatting_data1 is None:
                print("Analyzing formatting data for File 1...")
                self._cached_formatting_data1 = self._analyze_formatting_data(
                    self._cached_ws1, self._cached_blocks1
                )
                print(f"File 1: {len(self._cached_formatting_data1)} blocks with formatting data cached")
            return self._cached_formatting_data1
        else:
            if self._cached_formatting_data2 is None:
                print("Analyzing formatting data for File 2...")
                self._cached_formatting_data2 = self._analyze_formatting_data(
                    self._cached_ws2, self._cached_blocks2
                )
                print(f"File 2: {len(self._cached_formatting_data2)} blocks with formatting data cached")
            return self._cached_formatting_data2
    
    def _analyze_formatting_data(self, worksheet, blocks):
        """Analyze formatting data for all blocks at once."""
        formatting_data = []
        
        for i, block in enumerate(blocks):
            block_data = {
                'block_index': i,
                'block_rows': block,
                'ledger_accounts': [],
                'narration_short_codes': [],
                'amounts': {}
            }
            
            # Check each row in the block
            for row_idx in block:
                excel_row = row_idx + 10  # Convert to Excel row number
                
                if excel_row <= worksheet.max_row:
                    cell_c = worksheet.cell(row=excel_row, column=3)  # Column C
                    
                    # Check for ledger accounts (Bold but not italic)
                    if (cell_c.value and 
                        cell_c.font and 
                        cell_c.font.bold and 
                        not cell_c.font.italic):
                        
                        # Check if this is an interunit account
                        for full_account, short_code in self.interunit_loan_matcher.interunit_account_mapping.items():
                            if full_account.upper() in str(cell_c.value).upper():
                                block_data['ledger_accounts'].append({
                                    'full_account': full_account,
                                    'short_code': short_code,
                                    'cell_value': cell_c.value
                                })
                    
                    # Check for narration rows (Italic but not bold)
                    elif (cell_c.value and 
                          cell_c.font and 
                          not cell_c.font.bold and 
                          cell_c.font.italic):
                        
                        # Look for short codes in narration
                        for short_code in self.interunit_loan_matcher.interunit_account_mapping.values():
                            if short_code in str(cell_c.value):
                                block_data['narration_short_codes'].append({
                                    'short_code': short_code,
                                    'narration': cell_c.value
                                })
                    
                    # Check for amounts (Debit/Credit columns)
                    debit_cell = worksheet.cell(row=excel_row, column=8)  # Column H
                    credit_cell = worksheet.cell(row=excel_row, column=9)  # Column I
                    
                    if (debit_cell.value is not None and debit_cell.value != 0) or \
                       (credit_cell.value is not None and credit_cell.value != 0):
                        block_data['amounts'] = {
                            'debit': debit_cell.value if debit_cell.value else None,
                            'credit': credit_cell.value if credit_cell.value else None,
                            'row': row_idx
                        }
            
            if block_data['ledger_accounts'] or block_data['narration_short_codes']:
                formatting_data.append(block_data)
        
        return formatting_data
    
    def close_cached_workbooks(self):
        """Close cached workbooks to free memory."""
        if self._cached_wb1:
            self._cached_wb1.close()
            self._cached_wb1 = None
            self._cached_ws1 = None
        if self._cached_wb2:
            self._cached_wb2.close()
            self._cached_wb2 = None
            self._cached_ws2 = None
    
    def clear_all_caches(self):
        """Clear all caches to free memory."""
        # Clear workbook caches
        self.close_cached_workbooks()
        
        # Clear other caches
        self._cached_blocks1 = None
        self._cached_blocks2 = None
        self._cached_formatting_data1 = None
        self._cached_formatting_data2 = None
        self._cached_extracted_data = None
        
        # Clear amount caches
        self._amount_cache1.clear()
        self._amount_cache2.clear()
        self._block_header_cache1.clear()
        self._block_header_cache2.clear()
        
        # Clear additional performance caches
        self._block_header_cache.clear()
        self._description_row_cache.clear()
        self._narration_cache.clear()
        self._amount_cache.clear()
        
        print("All caches cleared for memory management")
    
    def create_output_files_optimized(self, all_matches, transactions1, transactions2, metadata1, metadata2):
        """Create output files using optimized batch operations."""
        print("Creating output files with optimized batch operations...")
        
        # Create output directory if it doesn't exist
        os.makedirs(OUTPUT_FOLDER, exist_ok=True)
        
        # Prepare data for batch insertion
        output_data = []
        
        for match in all_matches:
            file1_idx = match['File1_Index']
            file2_idx = match['File2_Index']
            match_type = match.get('Match_Type', 'Unknown')
            match_id = match.get('match_id', 'Unknown')
            
            # Get cached data for performance
            file1_header = self.get_cached_block_header_universal(file1_idx, transactions1)
            file2_header = self.get_cached_block_header_universal(file2_idx, transactions2)
            
            file1_desc = self.get_cached_description_row_universal(file1_idx, transactions1)
            file2_desc = self.get_cached_description_row_universal(file2_idx, transactions2)
            
            file1_narration = self.get_cached_narration(file1_desc, transactions1) if file1_desc else ""
            file2_narration = self.get_cached_narration(file2_desc, transactions2) if file2_desc else ""
            
            file1_amounts = self.get_cached_amounts_universal(file1_header, transactions1)
            file2_amounts = self.get_cached_amounts_universal(file2_header, transactions2)
            
            # Create audit info
            audit_info = self.create_audit_info(match)
            
            # Prepare row data
            row_data = [
                match_id,
                match_type,
                file1_idx,
                file2_idx,
                file1_narration,
                file2_narration,
                file1_amounts[0],  # debit
                file1_amounts[1],  # credit
                file2_amounts[0],  # debit
                file2_amounts[1],  # credit
                audit_info
            ]
            
            output_data.append(row_data)
        
        # Create DataFrame for batch operations
        output_df = pd.DataFrame(output_data, columns=[
            'Match_ID', 'Match_Type', 'File1_Index', 'File2_Index',
            'File1_Narration', 'File2_Narration', 'File1_Debit', 'File1_Credit',
            'File2_Debit', 'File2_Credit', 'Audit_Info'
        ])
        
        # Save to Excel with optimized formatting
        output_file = os.path.join(OUTPUT_FOLDER, f"matched_transactions{OUTPUT_SUFFIX}.xlsx")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write metadata
            metadata1.to_excel(writer, sheet_name='File1_Metadata', index=False, header=False)
            metadata2.to_excel(writer, sheet_name='File2_Metadata', index=False, header=False)
            
            # Write matches with optimized formatting
            output_df.to_excel(writer, sheet_name='Matches', index=False)
            
            # Get the worksheet for formatting
            worksheet = writer.sheets['Matches']
            
            # Apply formatting in batch
            self._apply_batch_formatting(worksheet, len(output_data))
        
        print(f"Output file created: {output_file}")
        return output_file
    
    def _apply_batch_formatting(self, worksheet, num_rows):
        """Apply formatting in batch for better performance."""
        # Set column widths
        column_widths = {
            'A': 9.00, 'B': 30.00, 'C': 12.00, 'D': 10.33, 'E': 60.00,
            'F': 5.00, 'G': 5.00, 'H': 12.78, 'I': 9.00, 'J': 13.78, 'K': 14.22
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Format amount columns in batch
        for row in range(2, num_rows + 2):  # Skip header row
            for col in ['H', 'I', 'J', 'K']:  # Amount columns
                cell = worksheet[f"{col}{row}"]
                if cell.value is not None:
                    cell.number_format = '#,##0.00'
    
    def create_optimized_arrays(self, lc_numbers1, lc_numbers2, po_numbers1, po_numbers2, 
                               usd_amounts1, usd_amounts2, interunit_accounts1, interunit_accounts2):
        """Create optimized numpy arrays for faster access in matching loops."""
        print("Creating optimized arrays for faster data access...")
        
        # Convert Series to numpy arrays for faster access
        self._lc_numbers1_array = lc_numbers1.values
        self._lc_numbers2_array = lc_numbers2.values
        self._po_numbers1_array = po_numbers1.values
        self._po_numbers2_array = po_numbers2.values
        self._usd_amounts1_array = usd_amounts1.values
        self._usd_amounts2_array = usd_amounts2.values
        self._interunit_accounts1_array = interunit_accounts1.values
        self._interunit_accounts2_array = interunit_accounts2.values
        
        print("Optimized arrays created successfully")
    
    def create_unmatched_indices_optimized(self, *match_lists):
        """Create unmatched indices efficiently without list concatenation."""
        matched_indices1 = set()
        matched_indices2 = set()
        
        # Process all match lists in one pass
        for match_list in match_lists:
            for match in match_list:
                if 'File1_Index' in match:
                    matched_indices1.add(match['File1_Index'])
                if 'File2_Index' in match:
                    matched_indices2.add(match['File2_Index'])
        
        return matched_indices1, matched_indices2
    
    def get_cached_block_header_universal(self, row_idx, transactions_df):
        """Get block header with universal caching for maximum performance."""
        cache_key = f"{id(transactions_df)}_{row_idx}"
        
        if cache_key not in self._block_header_cache:
            self._block_header_cache[cache_key] = self.block_identifier.find_transaction_block_header(row_idx, transactions_df)
        
        return self._block_header_cache[cache_key]
    
    def get_cached_description_row_universal(self, row_idx, transactions_df):
        """Get description row with universal caching for maximum performance."""
        cache_key = f"{id(transactions_df)}_{row_idx}"
        
        if cache_key not in self._description_row_cache:
            self._description_row_cache[cache_key] = self.block_identifier.find_description_row_in_block(row_idx, transactions_df)
        
        return self._description_row_cache[cache_key]
    
    def get_cached_narration(self, row_idx, transactions_df):
        """Get narration with caching for maximum performance."""
        cache_key = f"{id(transactions_df)}_{row_idx}"
        
        if cache_key not in self._narration_cache:
            narration = str(transactions_df.iloc[row_idx, 2]).strip()
            self._narration_cache[cache_key] = narration
        
        return self._narration_cache[cache_key]
    
    def get_cached_amounts_universal(self, row_idx, transactions_df):
        """Get amounts with universal caching for maximum performance."""
        cache_key = f"{id(transactions_df)}_{row_idx}"
        
        if cache_key not in self._amount_cache:
            row = transactions_df.iloc[row_idx]
            self._amount_cache[cache_key] = self.extract_amounts_from_strings(row)
        
        return self._amount_cache[cache_key]
    
    def precompute_all_block_data(self, transactions_df):
        """Precompute all block headers, description rows, and narrations for maximum performance."""
        print("Precomputing all block data for maximum performance...")
        
        for idx in range(len(transactions_df)):
            # Precompute block header
            self.get_cached_block_header_universal(idx, transactions_df)
            
            # Precompute description row
            self.get_cached_description_row_universal(idx, transactions_df)
            
            # Precompute narration
            self.get_cached_narration(idx, transactions_df)
            
            # Precompute amounts
            self.get_cached_amounts_universal(idx, transactions_df)
        
        print(f"Precomputed block data for {len(transactions_df)} rows")
    
    def get_optimized_lc_numbers(self, file_num, idx):
        """Get LC numbers using optimized array access."""
        if file_num == 1:
            return self._lc_numbers1_array[idx] if idx < len(self._lc_numbers1_array) else None
        else:
            return self._lc_numbers2_array[idx] if idx < len(self._lc_numbers2_array) else None
    
    def get_optimized_po_numbers(self, file_num, idx):
        """Get PO numbers using optimized array access."""
        if file_num == 1:
            return self._po_numbers1_array[idx] if idx < len(self._po_numbers1_array) else None
        else:
            return self._po_numbers2_array[idx] if idx < len(self._po_numbers2_array) else None
    
    def get_optimized_usd_amounts(self, file_num, idx):
        """Get USD amounts using optimized array access."""
        if file_num == 1:
            return self._usd_amounts1_array[idx] if idx < len(self._usd_amounts1_array) else None
        else:
            return self._usd_amounts2_array[idx] if idx < len(self._usd_amounts2_array) else None
    
    def get_optimized_interunit_accounts(self, file_num, idx):
        """Get interunit accounts using optimized array access."""
        if file_num == 1:
            return self._interunit_accounts1_array[idx] if idx < len(self._interunit_accounts1_array) else None
        else:
            return self._interunit_accounts2_array[idx] if idx < len(self._interunit_accounts2_array) else None
    
    def get_cached_block_header(self, idx, transactions_df, file_num):
        """Get transaction block header with caching for performance."""
        cache = self._block_header_cache1 if file_num == 1 else self._block_header_cache2
        
        if idx not in cache:
            cache[idx] = self.block_identifier.find_transaction_block_header(idx, transactions_df)
        
        return cache[idx]
    
    def get_cached_amounts(self, idx, transactions_df, file_num):
        """Get amounts with caching for performance."""
        cache = self._amount_cache1 if file_num == 1 else self._amount_cache2
        
        if idx not in cache:
            row = transactions_df.iloc[idx]
            cache[idx] = self.extract_amounts_from_strings(row)
        
        return cache[idx]
    
    def preprocess_all_amounts(self, transactions_df, file_num):
        """Preprocess all amounts for a file to populate cache."""
        print(f"Preprocessing amounts for File {file_num}...")
        cache = self._amount_cache1 if file_num == 1 else self._amount_cache2
        
        for idx in range(len(transactions_df)):
            if idx not in cache:
                row = transactions_df.iloc[idx]
                cache[idx] = self.extract_amounts_from_strings(row)
        
        print(f"Preprocessed {len(cache)} amounts for File {file_num}")
    
    def create_amount_index(self, transactions_df, file_num):
        """Create an index of transactions by amount for fast lookup."""
        print(f"Creating amount index for File {file_num}...")
        amount_index = {}
        cache = self._amount_cache1 if file_num == 1 else self._amount_cache2
        
        for idx in range(len(transactions_df)):
            if idx in cache:
                debit, credit = cache[idx]
                
                if debit > 0:  # Lender transaction
                    if debit not in amount_index:
                        amount_index[debit] = {'lenders': [], 'borrowers': []}
                    amount_index[debit]['lenders'].append(idx)
                
                if credit > 0:  # Borrower transaction
                    if credit not in amount_index:
                        amount_index[credit] = {'lenders': [], 'borrowers': []}
                    amount_index[credit]['borrowers'].append(idx)
        
        print(f"Created amount index with {len(amount_index)} unique amounts for File {file_num}")
        return amount_index
    
    def find_lc_matches_optimized(self, transactions1, transactions2, lc_numbers1, lc_numbers2, existing_matches=None, match_id_manager=None):
        """Optimized LC matching using amount pre-filtering."""
        matches = []
        
        print(f"\n=== OPTIMIZED LC MATCHING ===")
        print(f"Using amount pre-filtering for performance...")
        
        # Get all unique amounts that exist in both files
        common_amounts = set(self.amount_index1.keys()) & set(self.amount_index2.keys())
        print(f"Found {len(common_amounts)} common amounts between files")
        
        for amount in common_amounts:
            # Get all lender/borrower pairs for this amount
            file1_lenders = self.amount_index1[amount]['lenders']
            file1_borrowers = self.amount_index1[amount]['borrowers']
            file2_lenders = self.amount_index2[amount]['lenders']
            file2_borrowers = self.amount_index2[amount]['borrowers']
            
            # Check all possible lender-borrower combinations
            for lender_idx in file1_lenders:
                for borrower_idx in file2_borrowers:
                    lc1 = self.get_optimized_lc_numbers(1, lender_idx)
                    lc2 = self.get_optimized_lc_numbers(2, borrower_idx)
                    if lc1 and lc2 and lc1 == lc2:
                        # Found LC match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'LC',
                            'File1_Index': lender_idx,
                            'File2_Index': borrower_idx,
                            'LC_Number': lc1,
                            'Amount': amount
                        })
            
            # Check reverse combinations
            for lender_idx in file2_lenders:
                for borrower_idx in file1_borrowers:
                    lc1 = self.get_optimized_lc_numbers(1, borrower_idx)
                    lc2 = self.get_optimized_lc_numbers(2, lender_idx)
                    if lc1 and lc2 and lc1 == lc2:
                        # Found LC match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'LC',
                            'File1_Index': borrower_idx,
                            'File2_Index': lender_idx,
                            'LC_Number': lc1,
                            'Amount': amount
                        })
        
        print(f"Found {len(matches)} LC matches using optimized method")
        return matches
    
    def find_po_matches_optimized(self, transactions1, transactions2, po_numbers1, po_numbers2, existing_matches=None, match_id_manager=None):
        """Optimized PO matching using amount pre-filtering."""
        matches = []
        
        print(f"\n=== OPTIMIZED PO MATCHING ===")
        print(f"Using amount pre-filtering for performance...")
        
        # Get all unique amounts that exist in both files
        common_amounts = set(self.amount_index1.keys()) & set(self.amount_index2.keys())
        print(f"Found {len(common_amounts)} common amounts between files")
        
        for amount in common_amounts:
            # Get all lender/borrower pairs for this amount
            file1_lenders = self.amount_index1[amount]['lenders']
            file1_borrowers = self.amount_index1[amount]['borrowers']
            file2_lenders = self.amount_index2[amount]['lenders']
            file2_borrowers = self.amount_index2[amount]['borrowers']
            
            # Check all possible lender-borrower combinations
            for lender_idx in file1_lenders:
                for borrower_idx in file2_borrowers:
                    po1 = self.get_optimized_po_numbers(1, lender_idx)
                    po2 = self.get_optimized_po_numbers(2, borrower_idx)
                    if po1 and po2 and po1 == po2:
                        # Found PO match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'PO',
                            'File1_Index': lender_idx,
                            'File2_Index': borrower_idx,
                            'PO_Number': po1,
                            'Amount': amount
                        })
            
            # Check reverse combinations
            for lender_idx in file2_lenders:
                for borrower_idx in file1_borrowers:
                    po1 = self.get_optimized_po_numbers(1, borrower_idx)
                    po2 = self.get_optimized_po_numbers(2, lender_idx)
                    if po1 and po2 and po1 == po2:
                        # Found PO match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'PO',
                            'File1_Index': borrower_idx,
                            'File2_Index': lender_idx,
                            'PO_Number': po1,
                            'Amount': amount
                        })
        
        print(f"Found {len(matches)} PO matches using optimized method")
        return matches
    
    def find_usd_matches_optimized(self, transactions1, transactions2, usd_amounts1, usd_amounts2, existing_matches=None, match_id_manager=None):
        """Optimized USD matching using amount pre-filtering."""
        matches = []
        
        print(f"\n=== OPTIMIZED USD MATCHING ===")
        print(f"Using amount pre-filtering for performance...")
        
        # Get all unique amounts that exist in both files
        common_amounts = set(self.amount_index1.keys()) & set(self.amount_index2.keys())
        print(f"Found {len(common_amounts)} common amounts between files")
        
        for amount in common_amounts:
            # Get all lender/borrower pairs for this amount
            file1_lenders = self.amount_index1[amount]['lenders']
            file1_borrowers = self.amount_index1[amount]['borrowers']
            file2_lenders = self.amount_index2[amount]['lenders']
            file2_borrowers = self.amount_index2[amount]['borrowers']
            
            # Check all possible lender-borrower combinations
            for lender_idx in file1_lenders:
                for borrower_idx in file2_borrowers:
                    usd1 = self.get_optimized_usd_amounts(1, lender_idx)
                    usd2 = self.get_optimized_usd_amounts(2, borrower_idx)
                    if usd1 and usd2 and usd1 == usd2:
                        # Found USD match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'USD',
                            'File1_Index': lender_idx,
                            'File2_Index': borrower_idx,
                            'USD_Amount': usd1,
                            'Amount': amount
                        })
            
            # Check reverse combinations
            for lender_idx in file2_lenders:
                for borrower_idx in file1_borrowers:
                    usd1 = self.get_optimized_usd_amounts(1, borrower_idx)
                    usd2 = self.get_optimized_usd_amounts(2, lender_idx)
                    if usd1 and usd2 and usd1 == usd2:
                        # Found USD match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'USD',
                            'File1_Index': borrower_idx,
                            'File2_Index': lender_idx,
                            'USD_Amount': usd1,
                            'Amount': amount
                        })
        
        print(f"Found {len(matches)} USD matches using optimized method")
        return matches
    
    def find_narration_matches_optimized(self, transactions1, transactions2, existing_matches=None, match_id_manager=None):
        """Optimized Narration matching using amount pre-filtering."""
        matches = []
        
        print(f"\n=== OPTIMIZED NARRATION MATCHING ===")
        print(f"Using amount pre-filtering for performance...")
        
        # Get all unique amounts that exist in both files
        common_amounts = set(self.amount_index1.keys()) & set(self.amount_index2.keys())
        print(f"Found {len(common_amounts)} common amounts between files")
        
        for amount in common_amounts:
            # Get all lender/borrower pairs for this amount
            file1_lenders = self.amount_index1[amount]['lenders']
            file1_borrowers = self.amount_index1[amount]['borrowers']
            file2_lenders = self.amount_index2[amount]['lenders']
            file2_borrowers = self.amount_index2[amount]['borrowers']
            
            # Check all possible lender-borrower combinations
            for lender_idx in file1_lenders:
                for borrower_idx in file2_borrowers:
                    # Get narrations for comparison using cached method
                    lender_narration = self.get_cached_narration(lender_idx, transactions1)
                    borrower_narration = self.get_cached_narration(borrower_idx, transactions2)
                    
                    # Check for exact narration match
                    if (len(lender_narration) > 10 and len(borrower_narration) > 10 and 
                        lender_narration.lower() not in ['nan', 'none', ''] and
                        borrower_narration.lower() not in ['nan', 'none', ''] and
                        lender_narration == borrower_narration):
                        # Found narration match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'Narration',
                            'File1_Index': lender_idx,
                            'File2_Index': borrower_idx,
                            'Narration': lender_narration,
                            'Amount': amount
                        })
            
            # Check reverse combinations
            for lender_idx in file2_lenders:
                for borrower_idx in file1_borrowers:
                    # Get narrations for comparison using cached method
                    lender_narration = self.get_cached_narration(lender_idx, transactions2)
                    borrower_narration = self.get_cached_narration(borrower_idx, transactions1)
                    
                    # Check for exact narration match
                    if (len(lender_narration) > 10 and len(borrower_narration) > 10 and 
                        lender_narration.lower() not in ['nan', 'none', ''] and
                        borrower_narration.lower() not in ['nan', 'none', ''] and
                        lender_narration == borrower_narration):
                        # Found narration match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'Narration',
                            'File1_Index': borrower_idx,
                            'File2_Index': lender_idx,
                            'Narration': lender_narration,
                            'Amount': amount
                        })
        
        print(f"Found {len(matches)} Narration matches using optimized method")
        return matches
    
    def find_interunit_matches_optimized(self, transactions1, transactions2, interunit_accounts1, interunit_accounts2, existing_matches=None, match_id_manager=None):
        """Optimized Interunit matching using cached formatting data."""
        matches = []
        
        print(f"\n=== OPTIMIZED INTERUNIT MATCHING ===")
        print(f"Using cached formatting data for performance...")
        
        # Get cached formatting data
        file1_interunit_data = self.get_cached_formatting_data(1)
        file2_interunit_data = self.get_cached_formatting_data(2)
        
        print(f"File 1: {len(file1_interunit_data)} blocks with interunit data")
        print(f"File 2: {len(file2_interunit_data)} blocks with interunit data")
        
        # Look for cross-referenced matches using cached data
        print(f"\n--- Looking for cross-referenced matches ---")
        
        for block1 in file1_interunit_data:
            for block2 in file2_interunit_data:
                # Check if blocks have opposite transaction types (one debit, one credit)
                if (block1['amounts'] and block2['amounts'] and
                    ((block1['amounts']['debit'] and block2['amounts']['credit']) or
                     (block1['amounts']['credit'] and block2['amounts']['debit']))):
                    
                    # Check if amounts match exactly (NO TOLERANCE)
                    amount1 = block1['amounts']['debit'] if block1['amounts']['debit'] else block1['amounts']['credit']
                    amount2 = block2['amounts']['debit'] if block2['amounts']['debit'] else block2['amounts']['credit']
                    
                    if amount1 == amount2:
                        # Check for cross-referenced short codes
                        cross_reference_found = False
                        file1_narration_contains = None
                        file2_narration_contains = None
                        
                        # File 1's narration should contain File 2's short code
                        for narration1 in block1['narration_short_codes']:
                            for ledger2 in block2['ledger_accounts']:
                                if narration1['short_code'] == ledger2['short_code']:
                                    cross_reference_found = True
                                    file1_narration_contains = narration1['short_code']
                                    break
                            if cross_reference_found:
                                break
                        
                        # File 2's narration should contain File 1's short code
                        if cross_reference_found:
                            for narration2 in block2['narration_short_codes']:
                                for ledger1 in block1['ledger_accounts']:
                                    if narration2['short_code'] == ledger1['short_code']:
                                        file2_narration_contains = narration2['short_code']
                                        
                                        # We have a match!
                                        match = {
                                            'match_id': None,
                                            'Match_Type': 'Interunit',
                                            'Interunit_Account': f"{file1_narration_contains} <-> {file2_narration_contains}",
                                            'File1_Index': block1['amounts']['row'],
                                            'File2_Index': block2['amounts']['row'],
                                            'File1_Debit': block1['amounts']['debit'],
                                            'File1_Credit': block1['amounts']['credit'],
                                            'File2_Debit': block2['amounts']['debit'],
                                            'File2_Credit': block2['amounts']['credit'],
                                            'File1_Amount': amount1,
                                            'File2_Amount': amount1,
                                            'Amount': amount1
                                        }
                                        
                                        matches.append(match)
                                        print(f"  MATCH: Amount {amount1}")
                                        print(f"    Cross-reference: File 1 narration contains {file1_narration_contains}")
                                        print(f"    Cross-reference: File 2 narration contains {file2_narration_contains}")
                                        break
                                
                                if file2_narration_contains:
                                    break
        
        print(f"Found {len(matches)} Interunit matches using optimized method")
        return matches
    
    def find_aggregated_po_matches_optimized(self, transactions1, transactions2, po_numbers1, po_numbers2, existing_matches=None, match_id_manager=None):
        """Optimized Aggregated PO matching using amount pre-filtering."""
        matches = []
        
        print(f"\n=== OPTIMIZED AGGREGATED PO MATCHING ===")
        print(f"Using amount pre-filtering for performance...")
        
        # Get all unique amounts that exist in both files
        common_amounts = set(self.amount_index1.keys()) & set(self.amount_index2.keys())
        print(f"Found {len(common_amounts)} common amounts between files")
        
        # Import regex for PO pattern matching
        import re
        from .po_matching_logic import PO_PATTERN
        
        for amount in common_amounts:
            # Get all lender/borrower pairs for this amount
            file1_lenders = self.amount_index1[amount]['lenders']
            file1_borrowers = self.amount_index1[amount]['borrowers']
            file2_lenders = self.amount_index2[amount]['lenders']
            file2_borrowers = self.amount_index2[amount]['borrowers']
            
            # Check all possible lender-borrower combinations
            for lender_idx in file1_lenders:
                for borrower_idx in file2_borrowers:
                    # Get narrations for PO extraction using cached method
                    lender_narration = self.get_cached_narration(lender_idx, transactions1)
                    borrower_narration = self.get_cached_narration(borrower_idx, transactions2)
                    
                    # Extract PO numbers from narrations using compiled pattern
                    lender_pos = self._compiled_po_pattern.findall(lender_narration)
                    borrower_pos = self._compiled_po_pattern.findall(borrower_narration)
                    
                    # Check if lender has multiple POs and borrower has matching POs
                    if len(lender_pos) >= 2 and len(borrower_pos) >= 1:
                        # Check if all lender POs are present in borrower
                        if all(po in borrower_pos for po in lender_pos):
                            # Found aggregated PO match
                            matches.append({
                                'match_id': None,
                                'Match_Type': 'Aggregated_PO',
                                'File1_Index': lender_idx,
                                'File2_Index': borrower_idx,
                                'PO_Count': len(lender_pos),
                                'All_POs': lender_pos,
                                'Amount': amount
                            })
            
            # Check reverse combinations
            for lender_idx in file2_lenders:
                for borrower_idx in file1_borrowers:
                    # Get narrations for PO extraction using cached method
                    lender_narration = self.get_cached_narration(lender_idx, transactions2)
                    borrower_narration = self.get_cached_narration(borrower_idx, transactions1)
                    
                    # Extract PO numbers from narrations using compiled pattern
                    lender_pos = self._compiled_po_pattern.findall(lender_narration)
                    borrower_pos = self._compiled_po_pattern.findall(borrower_narration)
                    
                    # Check if lender has multiple POs and borrower has matching POs
                    if len(lender_pos) >= 2 and len(borrower_pos) >= 1:
                        # Check if all lender POs are present in borrower
                        if all(po in borrower_pos for po in lender_pos):
                            # Found aggregated PO match
                            matches.append({
                                'match_id': None,
                                'Match_Type': 'Aggregated_PO',
                                'File1_Index': borrower_idx,
                                'File2_Index': lender_idx,
                                'PO_Count': len(lender_pos),
                                'All_POs': lender_pos,
                                'Amount': amount
                            })
        
        print(f"Found {len(matches)} Aggregated PO matches using optimized method")
        return matches
    
    def extract_lc_numbers(self, description_series):
        """Extract LC numbers from transaction descriptions."""
        def extract_single_lc(description):
            if pd.isna(description):
                return None
            
            # Pattern for LC numbers: L/C-123/456, LC-123/456, or similar formats
            match = self._compiled_lc_pattern.search(str(description).upper())
            return match.group() if match else None
        
        return description_series.apply(extract_single_lc)
    
    def extract_po_numbers(self, description_series):
        """Extract PO numbers from transaction descriptions."""
        def extract_single_po(description):
            if pd.isna(description):
                return None
            
            # Pattern for PO numbers: XXX/PO/YYYY/M/NNNNN format
            match = self._compiled_po_pattern.search(str(description).upper())
            return match.group() if match else None
        
        return description_series.apply(extract_single_po)
    
    def extract_lc_numbers_from_narration(self, file_path):
        """Extract LC numbers from narration rows (regular text Column C - not bold, not italic) using openpyxl formatting."""
        lc_numbers = []
        lc_parent_rows = []
        
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row in range(9, ws.max_row + 1):  # Start from row 9 (after headers)
            desc_cell = ws.cell(row=row, column=3)
            
            # Check if this is a narration row (italic text Column C - not bold, but italic)
            is_narration = (desc_cell.value and 
                           desc_cell.font and 
                           not desc_cell.font.bold and 
                           desc_cell.font.italic)
            
            if is_narration:
                # This is a narration row, check for LC numbers
                narration_text = str(desc_cell.value)
                lc = self.extract_lc_numbers(pd.Series([narration_text])).iloc[0]
                
                if lc is not None:
                    # Found LC in narration row, need to find parent transaction row
                    parent_row = self.find_parent_transaction_row_with_formatting(ws, row)
                    if parent_row is not None:

                        lc_numbers.append(lc)
                        lc_parent_rows.append(parent_row)
                    else:

                        lc_numbers.append(None)
                        lc_parent_rows.append(None)
                else:
                    lc_numbers.append(None)
                    lc_parent_rows.append(None)
            else:
                lc_numbers.append(None)
                lc_parent_rows.append(None)
        
        wb.close()
        
        # Store parent row mapping for later use
        
        return pd.Series(lc_numbers)
    
    def extract_po_numbers_from_narration(self, file_path):
        """Extract PO numbers from narration rows (italic text Column C - not bold, but italic) using openpyxl formatting."""
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Create a list to store PO numbers for each row in the DataFrame
        # We need to align with the transactions DataFrame structure
        po_numbers = []
        po_parent_rows = []
        
        # First, get the transactions DataFrame to know how many rows we need
        transactions_df = self.read_complex_excel(file_path)[1]  # Get transactions part
        total_rows = len(transactions_df)
        
        # Initialize with None for all rows
        for _ in range(total_rows):
            po_numbers.append(None)
            po_parent_rows.append(None)
        
        # Now scan for PO numbers in narration rows and map them to DataFrame indices
        for excel_row in range(9, ws.max_row + 1):  # Excel rows start from 9
            desc_cell = ws.cell(row=excel_row, column=3)
            
            # Check if this is a narration row (italic text Column C - not bold, but italic)
            is_narration = (desc_cell.value and 
                           desc_cell.font and 
                           not desc_cell.font.bold and 
                           desc_cell.font.italic)
            
            if is_narration:
                # This is a narration row, check for PO numbers
                narration_text = str(desc_cell.value)
                po = self.extract_po_numbers(pd.Series([narration_text])).iloc[0]
                
                if po is not None:
                    # Found PO in narration row, need to find parent transaction row
                    parent_row = self.find_parent_transaction_row_with_formatting(ws, excel_row)
                    if parent_row is not None:
                        # Convert Excel row to DataFrame index
                        df_index = parent_row - 9  # Excel row 9 = DataFrame index 0
                        if 0 <= df_index < total_rows:
                            # print(f"DEBUG: PO {po} at Excel row {excel_row} -> DataFrame index {df_index}")
                            po_numbers[df_index] = po
                            po_parent_rows[df_index] = df_index
                        else:
                            # print(f"DEBUG: PO {po} at Excel row {excel_row} - INVALID DataFrame index {df_index}")
                            pass
                    else:

                        pass
        
        wb.close()
        
        # Store parent row mapping for later use
        
        return pd.Series(po_numbers)

    def load_workbooks_and_extract_data_optimized(self):
        """
        Extract all required data using cached workbooks and compiled regex patterns.
        This is the most optimized version that reuses cached data.
        """
        if self._cached_extracted_data is not None:
            print("Using cached extracted data...")
            return self._cached_extracted_data
        
        print("Extracting data using cached workbooks and compiled regex...")
        
        # Use cached workbooks (already loaded)
        ws1 = self._cached_ws1
        ws2 = self._cached_ws2
        
        # Extract all data from File 1
        lc_numbers1 = []
        po_numbers1 = []
        usd_amounts1 = []
        interunit_accounts1 = []
        
        # Process File 1 rows 9 onwards (optimized with compiled regex)
        for row in range(9, ws1.max_row + 1):
            narration = ws1.cell(row=row, column=3).value  # Column C is narration
            if narration:
                # Cache upper case conversion
                narration_upper = str(narration).upper()
                
                # Extract LC numbers using compiled pattern
                lc_matches = self._compiled_lc_pattern.findall(narration_upper)
                if lc_matches:
                    lc_numbers1.append((row, lc_matches[0]))
                
                # Extract PO numbers using compiled pattern
                po_matches = self._compiled_po_pattern.findall(narration_upper)
                if po_matches:
                    po_numbers1.append((row, po_matches[0]))
                
                # Extract USD amounts using compiled pattern
                usd_matches = self._compiled_usd_pattern.findall(narration_upper)
                if usd_matches:
                    usd_amounts1.append((row, usd_matches[0]))
                
                # Extract interunit accounts using compiled pattern
                interunit_matches = self._compiled_interunit_pattern.findall(narration_upper)
                if interunit_matches:
                    interunit_accounts1.append((row, f"{interunit_matches[0][0]}#{interunit_matches[0][1]}"))
        
        # Convert to Series with proper indexing (matching original logic)
        # Create Series with same length as transactions DataFrame, initialized with None
        total_rows1 = len(self.transactions1)
        lc_numbers1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        po_numbers1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        usd_amounts1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        interunit_accounts1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        
        # Now populate the found items at their correct DataFrame indices
        for row, lc_num in lc_numbers1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                lc_numbers1_series.iloc[df_index] = lc_num
        
        for row, po_num in po_numbers1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                po_numbers1_series.iloc[df_index] = po_num
        
        for row, usd_amount in usd_amounts1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                usd_amounts1_series.iloc[df_index] = usd_amount
        
        for row, account in interunit_accounts1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                interunit_accounts1_series.iloc[df_index] = account
        
        # Extract all data from File 2
        lc_numbers2 = []
        po_numbers2 = []
        usd_amounts2 = []
        interunit_accounts2 = []
        
        # Process File 2 rows 9 onwards (optimized with compiled regex)
        for row in range(9, ws2.max_row + 1):
            narration = ws2.cell(row=row, column=3).value  # Column C is narration
            if narration:
                # Cache upper case conversion
                narration_upper = str(narration).upper()
                
                # Extract LC numbers using compiled pattern
                lc_matches = self._compiled_lc_pattern.findall(narration_upper)
                if lc_matches:
                    lc_numbers2.append((row, lc_matches[0]))
                
                # Extract PO numbers using compiled pattern
                po_matches = self._compiled_po_pattern.findall(narration_upper)
                if po_matches:
                    po_numbers2.append((row, po_matches[0]))
                
                # Extract USD amounts using compiled pattern
                usd_matches = self._compiled_usd_pattern.findall(narration_upper)
                if usd_matches:
                    usd_amounts2.append((row, usd_matches[0]))
                
                # Extract interunit accounts using compiled pattern
                interunit_matches = self._compiled_interunit_pattern.findall(narration_upper)
                if interunit_matches:
                    interunit_accounts2.append((row, f"{interunit_matches[0][0]}#{interunit_matches[0][1]}"))
        
        # Convert to Series with proper indexing (matching original logic)
        # Create Series with same length as transactions DataFrame, initialized with None
        total_rows2 = len(self.transactions2)
        lc_numbers2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        po_numbers2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        usd_amounts2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        interunit_accounts2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        
        # Now populate the found items at their correct DataFrame indices
        for row, lc_num in lc_numbers2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                lc_numbers2_series.iloc[df_index] = lc_num
        
        for row, po_num in po_numbers2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                po_numbers2_series.iloc[df_index] = po_num
        
        for row, usd_amount in usd_amounts2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                usd_amounts2_series.iloc[df_index] = usd_amount
        
        for row, account in interunit_accounts2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                interunit_accounts2_series.iloc[df_index] = account
        
        print(f"Data extraction complete:")
        print(f"  File 1: {len(lc_numbers1)} LC, {len(po_numbers1)} PO, {len(usd_amounts1)} USD, {len(interunit_accounts1)} Interunit")
        print(f"  File 2: {len(lc_numbers2)} LC, {len(po_numbers2)} PO, {len(usd_amounts2)} USD, {len(interunit_accounts2)} Interunit")
        
        # Cache the extracted data
        self._cached_extracted_data = {
            'lc_numbers1': lc_numbers1_series,
            'po_numbers1': po_numbers1_series,
            'usd_amounts1': usd_amounts1_series,
            'interunit_accounts1': interunit_accounts1_series,
            'lc_numbers2': lc_numbers2_series,
            'po_numbers2': po_numbers2_series,
            'usd_amounts2': usd_amounts2_series,
            'interunit_accounts2': interunit_accounts2_series
        }
        
        return self._cached_extracted_data

    def process_files(self):
        """Process both files and prepare for matching with performance optimizations."""
        print("Reading Pole Book STEEL.xlsx...")
        self.metadata1, self.transactions1 = self.read_complex_excel(self.file1_path)
        
        print("Reading Steel Book POLE.xlsx...")
        self.metadata2, self.transactions2 = self.read_complex_excel(self.file2_path)
        
        print(f"File 1: {len(self.transactions1)} rows")
        print(f"File 2: {len(self.transactions2)} rows")
        
        # Load and cache workbooks once for all operations
        self.load_and_cache_workbooks()
        
        # Preprocess amounts for performance optimization
        print("Preprocessing amounts for performance...")
        self.preprocess_all_amounts(self.transactions1, 1)
        self.preprocess_all_amounts(self.transactions2, 2)
        
        # Create amount indexes for fast lookup
        print("Creating amount indexes for fast lookup...")
        self.amount_index1 = self.create_amount_index(self.transactions1, 1)
        self.amount_index2 = self.create_amount_index(self.transactions2, 2)

        print(f"File 1 columns: {list(self.transactions1.columns)}")
        print(f"File 2 columns: {list(self.transactions2.columns)}")
        
        # Find the description column (should be the 3rd column, index 2)
        # Let's check what's actually in the columns
        print(f"File 1 first row: {list(self.transactions1.iloc[0, :])}")
        
        # Extract all data using cached workbooks and compiled regex
        print("Extracting all data using optimized methods...")
        extracted_data = self.load_workbooks_and_extract_data_optimized()
        
        # Extract LC numbers from both files
        lc_numbers1 = extracted_data['lc_numbers1']
        lc_numbers2 = extracted_data['lc_numbers2']
        
        # Extract PO numbers from both files
        po_numbers1 = extracted_data['po_numbers1']
        po_numbers2 = extracted_data['po_numbers2']
        
        # Extract interunit loan accounts from both files
        interunit_accounts1 = extracted_data['interunit_accounts1']
        interunit_accounts2 = extracted_data['interunit_accounts2']
        
        # Extract USD amounts from both files
        usd_amounts1 = extracted_data['usd_amounts1']
        usd_amounts2 = extracted_data['usd_amounts2']
        
        # Get cached transaction blocks (identified once, reused everywhere)
        print("Getting cached transaction blocks...")
        blocks1 = self.get_cached_blocks(1)
        blocks2 = self.get_cached_blocks(2)
        
        print(f"File 1: {len(blocks1)} transaction blocks")
        print(f"File 2: {len(blocks2)} transaction blocks")
        
        # Create optimized arrays for faster data access
        self.create_optimized_arrays(
            lc_numbers1, lc_numbers2, po_numbers1, po_numbers2,
            usd_amounts1, usd_amounts2, interunit_accounts1, interunit_accounts2
        )
        
        # Selective precomputation based on file size
        total_rows = len(self.transactions1) + len(self.transactions2)
        if total_rows > 2000:  # Only precompute for large files
            print("Large files detected - precomputing block data for maximum performance...")
            self.precompute_all_block_data(self.transactions1)
            self.precompute_all_block_data(self.transactions2)
        else:
            print("Small files detected - using on-demand caching for optimal performance...")
        
        return self.transactions1, self.transactions2, blocks1, blocks2, lc_numbers1, lc_numbers2, po_numbers1, po_numbers2, interunit_accounts1, interunit_accounts2, usd_amounts1, usd_amounts2
    
    def find_potential_matches(self):
        """Find potential LC, PO, Interunit, and USD matches between the two files (sequential approach)."""

        transactions1, transactions2, _, _, lc_numbers1, lc_numbers2, po_numbers1, po_numbers2, interunit_accounts1, interunit_accounts2, usd_amounts1, usd_amounts2 = self.process_files()
        
        print("\n" + "="*60)
        print("STEP 1: NARRATION MATCHING (HIGHEST PRIORITY - Most reliable)")
        print("="*60)
        
        # ARCHITECTURAL FIX: Collect all matches first, then assign sequential Match IDs
        print(f"\nMATCH ID SYSTEM: Post-processing sequential assignment")
        print(f"Expected sequence: M001, M002, M003... assigned after all matches are found")
        
        # Step 1: Find Narration matches (HIGHEST PRIORITY - Most reliable)
        print(f"\nSTEP 1: NARRATION MATCHING")
        narration_matches = self.find_narration_matches_optimized(
            transactions1, transactions2, {}, None
        )
        
        # Step 2: Find LC matches on UNMATCHED records
        print("\n" + "="*60)
        print("STEP 2: LC MATCHING (ON UNMATCHED RECORDS)")
        print("="*60)
        
        # Create masks for unmatched records (after Narration matching) - OPTIMIZED
        narration_matched_indices1, narration_matched_indices2 = self.create_unmatched_indices_optimized(narration_matches)
        
        # Filter LC numbers to only unmatched records
        lc_numbers1_unmatched = lc_numbers1.copy()
        lc_numbers2_unmatched = lc_numbers2.copy()
        
        # Mark matched records as None in LC numbers
        for idx in narration_matched_indices1:
            if idx < len(lc_numbers1_unmatched):
                lc_numbers1_unmatched.iloc[idx] = None
        
        for idx in narration_matched_indices2:
            if idx < len(lc_numbers2_unmatched):
                lc_numbers2_unmatched.iloc[idx] = None
        
        print(f"File 1: {len(lc_numbers1_unmatched[lc_numbers1_unmatched.notna()])} unmatched LC numbers")
        print(f"File 2: {len(lc_numbers2_unmatched[lc_numbers2_unmatched.notna()])} unmatched LC numbers")
        
        print(f"\nSTEP 2: LC MATCHING")
        # Use optimized LC matching for better performance
        lc_matches = self.find_lc_matches_optimized(
            transactions1, transactions2, lc_numbers1_unmatched, lc_numbers2_unmatched,
            {}, None
        )
        
        # print(f"\nLC Matching Results: {len(lc_matches)} matches found")
        
        # Step 3: Find PO matches on UNMATCHED records
        print("\n" + "="*60)
        print("STEP 3: PO MATCHING (ON UNMATCHED RECORDS)")
        print("="*60)
        
        # Create masks for unmatched records (after Narration and LC matching) - OPTIMIZED
        narration_lc_matched_indices1, narration_lc_matched_indices2 = self.create_unmatched_indices_optimized(narration_matches, lc_matches)
        
        # Filter PO numbers to only unmatched records
        po_numbers1_unmatched = po_numbers1.copy()
        po_numbers2_unmatched = po_numbers2.copy()
        
        # Mark matched records as None in PO numbers
        for idx in narration_lc_matched_indices1:
            if idx < len(po_numbers1_unmatched):
                po_numbers1_unmatched.iloc[idx] = None
        
        for idx in narration_lc_matched_indices2:
            if idx < len(po_numbers2_unmatched):
                po_numbers2_unmatched.iloc[idx] = None
        
        print(f"File 1: {len(po_numbers1_unmatched[po_numbers1_unmatched.notna()])} unmatched PO numbers")
        print(f"File 2: {len(po_numbers2_unmatched[po_numbers2_unmatched.notna()])} unmatched PO numbers")
        
        # Find PO matches on unmatched records with shared state
        print(f"\nSTEP 3: PO MATCHING")
        po_matches = self.find_po_matches_optimized(
            transactions1, transactions2, po_numbers1_unmatched, po_numbers2_unmatched,
            {}, None
        )
        
        # print(f"\nPO Matching Results: {len(po_matches)} matches found")
        
        # Step 4: Find Interunit Loan matches on UNMATCHED records
        print("\n" + "="*60)
        print("STEP 4: INTERUNIT LOAN MATCHING (ON UNMATCHED RECORDS)")
        print("="*60)
        
        # Create masks for unmatched records (after Narration, LC, and PO matching) - OPTIMIZED
        narration_lc_po_matched_indices1, narration_lc_po_matched_indices2 = self.create_unmatched_indices_optimized(narration_matches, lc_matches, po_matches)
        
        # Filter interunit accounts to only unmatched records
        interunit_accounts1_unmatched = interunit_accounts1.copy()
        interunit_accounts2_unmatched = interunit_accounts2.copy()
        
        # Mark matched records as None in interunit accounts
        for idx in narration_lc_po_matched_indices1:
            if idx < len(interunit_accounts1_unmatched):
                interunit_accounts1_unmatched.iloc[idx] = None
        
        for idx in narration_lc_po_matched_indices2:
            if idx < len(interunit_accounts2_unmatched):
                interunit_accounts2_unmatched.iloc[idx] = None
        
        print(f"File 1: {len(interunit_accounts1_unmatched[interunit_accounts1_unmatched.notna()])} unmatched interunit accounts")
        print(f"File 2: {len(interunit_accounts2_unmatched[interunit_accounts2_unmatched.notna()])} unmatched interunit accounts")
        
        # Find interunit loan matches on unmatched records with shared state
        print(f"\nSTEP 4: INTERUNIT MATCHING")
        interunit_matches = self.find_interunit_matches_optimized(
            transactions1, transactions2, interunit_accounts1_unmatched, interunit_accounts2_unmatched,
            {}, None
        )
        
        # print(f"\nInterunit Loan Matching Results: {len(interunit_matches)} matches found")
        
        # Step 5: Find USD matches on UNMATCHED records
        print("\n" + "="*60)
        print("STEP 5: USD MATCHING (ON UNMATCHED RECORDS)")
        print("="*60)
        
        # Create masks for unmatched records (after Narration, LC, PO, and Interunit matching) - OPTIMIZED
        narration_lc_po_interunit_matched_indices1, narration_lc_po_interunit_matched_indices2 = self.create_unmatched_indices_optimized(narration_matches, lc_matches, po_matches, interunit_matches)
        
        # Filter USD amounts to only unmatched records
        usd_amounts1_unmatched = usd_amounts1.copy()
        usd_amounts2_unmatched = usd_amounts2.copy()
        
        # Mark matched records as None in USD amounts
        for idx in narration_lc_po_interunit_matched_indices1:
            if idx < len(usd_amounts1_unmatched):
                usd_amounts1_unmatched.iloc[idx] = None
        
        for idx in narration_lc_po_interunit_matched_indices2:
            if idx < len(usd_amounts2_unmatched):
                usd_amounts2_unmatched.iloc[idx] = None
        
        print(f"File 1: {len(usd_amounts1_unmatched[usd_amounts1_unmatched.notna()])} unmatched USD amounts")
        print(f"File 2: {len(usd_amounts2_unmatched[usd_amounts2_unmatched.notna()])} unmatched USD amounts")
        
        # Find USD matches on unmatched records with shared state
        print(f"\nSTEP 5: USD MATCHING")
        usd_matches = self.find_usd_matches_optimized(
            transactions1, transactions2, usd_amounts1_unmatched, usd_amounts2_unmatched,
            {}, None
        )
        
        # print(f"\nUSD Matching Results: {len(usd_matches)} matches found")
        
        # Step 6: Find Aggregated PO matches on UNMATCHED records
        # COMMENTED OUT - One-to-many PO matches not working
        print("\n" + "="*60)
        print("STEP 6: AGGREGATED PO MATCHING - DISABLED")
        print("="*60)
        print("Aggregated PO matching has been commented out due to issues")
        
        # # Create masks for unmatched records (after Narration, LC, PO, Interunit, and USD matching)
        # narration_lc_po_interunit_usd_matched_indices1 = set()
        # narration_lc_po_interunit_usd_matched_indices2 = set()
        # 
        # for match in narration_matches + lc_matches + po_matches + interunit_matches + usd_matches:
        #     narration_lc_po_interunit_usd_matched_indices1.add(match['File1_Index'])
        #     narration_lc_po_interunit_usd_matched_indices2.add(match['File2_Index'])
        # 
        # # Filter PO numbers to only unmatched records
        # po_numbers1_unmatched_for_aggregated = po_numbers1.copy()
        # po_numbers2_unmatched_for_aggregated = po_numbers2.copy()
        # 
        # # Mark matched records as None in PO numbers
        # for idx in narration_lc_po_interunit_usd_matched_indices1:
        #     if idx < len(po_numbers1_unmatched_for_aggregated):
        #         po_numbers1_unmatched_for_aggregated.iloc[idx] = None
        # 
        # for idx in narration_lc_po_interunit_usd_matched_indices2:
        #     if idx < len(po_numbers2_unmatched_for_aggregated):
        #         po_numbers2_unmatched_for_aggregated.iloc[idx] = None
        # 
        # print(f"File 1: {len(po_numbers1_unmatched_for_aggregated[po_numbers1_unmatched_for_aggregated.notna()])} unmatched PO numbers for aggregated matching")
        # print(f"File 2: {len(po_numbers2_unmatched_for_aggregated[po_numbers2_unmatched_for_aggregated.notna()])} unmatched PO numbers for aggregated matching")
        # 
        # # Find aggregated PO matches on unmatched records with shared state
        # print(f"\nSTEP 6: AGGREGATED PO MATCHING")
        # aggregated_po_matches = self.find_aggregated_po_matches_optimized(
        #     transactions1, transactions2, po_numbers1_unmatched_for_aggregated, po_numbers2_unmatched_for_aggregated,
        #     {}, None
        # )
        # 
        # # print(f"\nAggregated PO Matching Results: {len(aggregated_po_matches)} matches found")
        
        # Set aggregated_po_matches to empty list since it's disabled
        aggregated_po_matches = []
        
        # Combine all matches
        all_matches = narration_matches + lc_matches + po_matches + interunit_matches + usd_matches + aggregated_po_matches
        
        # ARCHITECTURAL FIX: Assign sequential Match IDs to all matches
        print(f"\n=== ASSIGNING SEQUENTIAL MATCH IDs ===")
        print(f"Total matches found: {len(all_matches)}")

        
        # Initialize Match ID counter
        match_counter = 1
        
        # Assign sequential Match IDs to all matches
        for i, match in enumerate(all_matches):
            match_id = f"M{match_counter:03d}"  # Format as M001, M002, M003, etc.
            old_match_id = match.get('match_id', 'None')
            match['match_id'] = match_id
            match_counter += 1
            print(f"Match {i+1}: Assigned {match_id} to {match.get('Match_Type', 'Unknown')} match (was {old_match_id})")
        
        print(f"Assigned {len(all_matches)} sequential Match IDs (M001 to M{match_counter-1:03d})")
        
        # Sort matches by the newly assigned sequential Match IDs
        all_matches.sort(key=lambda x: x['match_id'])
        print(f"Sorted matches by sequential Match IDs")
        
        print(f"\n" + "="*60)
        print("FINAL MATCH SUMMARY")
        print("="*60)
        print(f"Total matches found: {len(all_matches)}")
        print(f"Match IDs assigned: M001 to M{match_counter-1:03d}")
        
        print("="*60)
        print("FINAL RESULTS")
        print("="*60)
        print(f"Total Matches: {len(all_matches)}")
        print(f"  - Narration Matches: {len(narration_matches)} (HIGHEST PRIORITY)")
        print(f"  - LC Matches: {len(lc_matches)}")
        print(f"  - PO Matches: {len(po_matches)}")
        print(f"  - Interunit Loan Matches: {len(interunit_matches)}")
        print(f"  - USD Matches: {len(usd_matches)}")
        print(f"  - Aggregated PO Matches: {len(aggregated_po_matches)} (DISABLED)")
        
        # Clean up all caches to free memory
        self.clear_all_caches()
        
        return all_matches
    
    def find_parent_transaction_row_with_formatting(self, ws, current_row):
        """Find the parent transaction row for a narration row using openpyxl formatting."""
        # Look backwards from current row to find the most recent transaction block header
        for row_idx in range(current_row, 8, -1):  # Start from current_row, go back to row 9
            date_cell = ws.cell(row=row_idx, column=1)
            particulars_cell = ws.cell(row=row_idx, column=2)
            desc_cell = ws.cell(row=row_idx, column=3)
            
            # Check if this is a transaction block header (Date + Dr/Cr + BOLD Col C)
            has_date = date_cell.value is not None
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            has_bold_desc = desc_cell.font and desc_cell.font.bold
            
            if has_date and has_dr_cr and has_bold_desc:
                return row_idx
        
        return None
    
    def create_audit_info(self, match):
        """Create audit info in clean, readable plaintext format for LC, PO, Interunit, and USD matches."""
        # Determine match type and create appropriate audit info
        if 'Match_Type' in match:
            # Use explicit match type if available
            match_type = match['Match_Type']
            amount = match.get('File1_Amount', match.get('File2_Amount', 0))
            
            if match_type == 'Narration':
                audit_info = f"Narration Match\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'LC':
                lc_number = match.get('LC_Number', 'Unknown')
                audit_info = f"LC Match: {lc_number}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'PO':
                po_number = match.get('PO_Number', 'Unknown')
                audit_info = f"PO Match: {po_number}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'Interunit':
                interunit_account = match.get('Interunit_Account', 'Unknown')
                audit_info = f"Interunit Loan Match: {interunit_account}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'USD':
                usd_amount = match.get('USD_Amount', 'Unknown')
                audit_info = f"USD Match: {usd_amount}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'Aggregated_PO':
                po_count = match.get('PO_Count', 'Unknown')
                all_pos = match.get('All_POs', [])
                po_list = ', '.join(all_pos[:5])  # Show first 5 POs
                if len(all_pos) > 5:
                    po_list += f" ... and {len(all_pos) - 5} more"
                audit_info = f"Aggregated PO Match: {po_count} POs\nPOs: {po_list}\nLender Amount: {amount:.2f}\nTotal Borrower Amount: {amount:.2f}"
            else:
                audit_info = f"{match_type} Match\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
        else:
            # Fallback to old logic for backward compatibility
            if 'LC_Number' in match and match['LC_Number']:
                # This is an LC match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"LC Match: {match['LC_Number']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif 'PO_Number' in match and match['PO_Number']:
                # This is a PO match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"PO Match: {match['PO_Number']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif 'Interunit_Account' in match and match['Interunit_Account']:
                # This is an Interunit Loan match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"Interunit Loan Match: {match['Interunit_Account']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif 'USD_Amount' in match and match['USD_Amount']:
                # This is a USD match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"USD Match: {match['USD_Amount']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            else:
                # Fallback for unknown match type
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"Unknown Match Type\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
        
        return audit_info

    def _format_amount_columns(self, worksheet):
        """Format amount columns (Debit and Credit) to prevent scientific notation."""
        # Debit column (J) and Credit column (K) - after adding Match ID and Audit Info
        debit_col = 9  # Column J (0-indexed)
        credit_col = 10  # Column K (0-indexed)
        
        # Format all data rows (starting from row 9)
        for row in range(9, worksheet.max_row + 1):
            try:
                # Format Debit column
                debit_cell = worksheet.cell(row=row, column=debit_col + 1)  # openpyxl uses 1-indexed
                if debit_cell.value is not None and debit_cell.value != '':
                    debit_cell.number_format = '#,##0.00'
                
                # Format Credit column
                credit_cell = worksheet.cell(row=row, column=credit_col + 1)  # openpyxl uses 1-indexed
                if credit_cell.value is not None and credit_cell.value != '':
                    credit_cell.number_format = '#,##0.00'
                    
            except Exception as e:
                print(f"Error formatting amount columns for row {row}: {e}")

    def _set_column_widths(self, worksheet):
        """Set column widths for the worksheet"""
        worksheet.column_dimensions['A'].width = 9.00
        worksheet.column_dimensions['B'].width = 30.00
        worksheet.column_dimensions['C'].width = 12.00
        worksheet.column_dimensions['D'].width = 10.33
        worksheet.column_dimensions['E'].width = 60.00
        worksheet.column_dimensions['F'].width = 5.00
        worksheet.column_dimensions['G'].width = 5.00
        worksheet.column_dimensions['H'].width = 12.78
        worksheet.column_dimensions['I'].width = 9.00
        worksheet.column_dimensions['J'].width = 13.78
        worksheet.column_dimensions['K'].width = 14.22
        worksheet.column_dimensions['L'].width = 11.22

    def _apply_top_alignment(self, worksheet):
        """Apply top alignment and text wrapping to ALL cells in the worksheet."""
        print(f"Setting top alignment for {worksheet.max_row} rows × {worksheet.max_column} columns...")
        
        for row in range(1, worksheet.max_row + 1):  # ALL rows from 1 to max
            for col in range(1, worksheet.max_column + 1):  # ALL columns
                try:
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Always create a new alignment object to avoid style conflicts
                    new_alignment = Alignment(vertical='top')
                    
                    # Enable text wrapping for columns B (Audit Info) and E (Description)
                    if col in [2, 5]:  # Columns B and E
                        new_alignment.wrap_text = True
                    
                    # Apply the new alignment (this overwrites any existing alignment)
                    cell.alignment = new_alignment
                        
                except Exception as e:
                    print(f"Error setting alignment for row {row}, col {col}: {e}")
                    # Continue with next cell instead of stopping
                    continue
        
        print(f"Top alignment applied successfully!")

    def _apply_filters_to_header(self, worksheet):
        """Apply filters to the header row (Row 9) for easy data filtering and sorting."""
        try:
            # Apply filters to Row 9 (header row)
            # Note: openpyxl uses 1-based indexing, so Row 9 is actually row 9
            worksheet.auto_filter.ref = f"A9:L9"
            print(f"Filters applied to header row (Row 9) successfully!")
        except Exception as e:
            print(f"Error applying filters to header row: {e}")
    
    def _apply_alternating_background_colors(self, worksheet, file_matched_df):
        """Apply alternating background colors to matched transaction blocks."""
        try:
            from openpyxl.styles import PatternFill
            
            # Define two alternating colors
            color1 = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Very light blue
            color2 = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Very light lemon yellow
            
            # Get all rows with Match IDs
            match_id_column = file_matched_df.iloc[:, 0]  # First column (Match ID)
            populated_rows = match_id_column.notna()
            
            if not populated_rows.any():
                print("No matched rows found for background coloring")
                return
            
            # Get unique Match IDs in order they appear
            unique_match_ids = []
            seen_ids = set()
            for _, match_id in enumerate(match_id_column):
                if pd.notna(match_id) and match_id not in seen_ids:
                    unique_match_ids.append(match_id)
                    seen_ids.add(match_id)
            
            print(f"Applying alternating background colors to {len(unique_match_ids)} matched transaction blocks")
            
            # Apply alternating colors to each Match ID block
            for block_index, match_id in enumerate(unique_match_ids):
                # Choose color based on block index (alternating)
                color = color1 if block_index % 2 == 0 else color2
                
                # Find all rows with this Match ID
                block_rows = file_matched_df[file_matched_df.iloc[:, 0] == match_id].index
                
                # Apply color to all rows in this block
                for df_row_idx in block_rows:
                    excel_row = df_row_idx + 10  # Convert DataFrame index to Excel row (metadata + header offset)
                    
                    # Color all columns in this row
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=excel_row, column=col)
                        cell.fill = color
                
                print(f"  Block {match_id}: Applied {'Color 1' if block_index % 2 == 0 else 'Color 2'} to {len(block_rows)} rows")
            
            print("Background colors applied successfully!")
            
        except Exception as e:
            print(f"Error applying background colors: {e}")

    def _format_output_file_transaction_blocks(self, worksheet):
        """Format output file transaction blocks: make ledger text bold, narration italic, and Entered By person's name bold+italic."""
        try:
            from openpyxl.styles import Font
            
            print("Formatting output file transaction blocks...")
            
            # Create fonts for different formatting
            bold_font = Font(bold=True)
            italic_font = Font(italic=True)
            bold_italic_font = Font(bold=True, italic=True)
            
            # Process all rows starting from row 10 (after metadata and header)
            for row in range(10, worksheet.max_row + 1):
                # Check if this row is the end of a transaction block
                cell_d = worksheet.cell(row=row, column=4)  # Column D (Particulars)
                cell_e = worksheet.cell(row=row, column=5)  # Column E (Description)
                
                # Check if this row contains "Entered By :" in Column D
                if (cell_d.value and 
                    isinstance(cell_d.value, str) and 
                    "Entered By :" in str(cell_d.value)):
                    
                    # This is the end of a transaction block
                    # Make the Entered By person's name bold and italic
                    if cell_e.value:
                        cell_e.font = bold_italic_font
                        print(f"  Row {row}: Made Entered By person's name bold+italic: '{str(cell_e.value)[:50]}...'")
                    
                    # The row above this contains narration text
                    narration_row = row - 1
                    
                    if narration_row >= 10:  # Ensure we don't go below row 10
                        narration_cell_e = worksheet.cell(row=narration_row, column=5)  # Column E
                        
                        # Make the narration text italic
                        if narration_cell_e.value:
                            narration_cell_e.font = italic_font
                            print(f"  Row {narration_row}: Made narration text italic: '{str(narration_cell_e.value)[:50]}...'")
                        
                        # Now find the transaction block start and make all ledger text bold
                        # Look backwards from narration row to find block start
                        for ledger_row in range(narration_row - 1, 9, -1):  # Go back from narration to row 10
                            # Check if this is a block start row
                            date_cell = worksheet.cell(row=ledger_row, column=3)  # Column C (Date)
                            particulars_cell = worksheet.cell(row=ledger_row, column=4)  # Column D (Particulars)
                            vch_type_cell = worksheet.cell(row=ledger_row, column=8)  # Column H (Vch Type)
                            vch_no_cell = worksheet.cell(row=ledger_row, column=9)  # Column I (Vch No)
                            
                            # Check if this is a block start (has date, Dr/Cr, Vch Type, Vch No)
                            is_block_start = (date_cell.value and 
                                            particulars_cell.value and 
                                            str(particulars_cell.value).strip() in ['Dr', 'Cr'] and
                                            vch_type_cell.value and 
                                            vch_no_cell.value)
                            
                            if is_block_start:
                                 # Found block start, now make all rows from here to narration bold
                                 for bold_row in range(ledger_row, narration_row):
                                     bold_cell_e = worksheet.cell(row=bold_row, column=5)  # Column E
                                     if bold_cell_e.value:
                                         bold_cell_e.font = bold_font
                                         print(f"  Row {bold_row}: Made ledger text bold: '{str(bold_cell_e.value)[:50]}...'")
                                 
                                 # Also make Column H (Vch Type) bold in the first row of the transaction block
                                 vch_type_cell = worksheet.cell(row=ledger_row, column=8)  # Column H (Vch Type)
                                 if vch_type_cell.value:
                                     vch_type_cell.font = bold_font
                                     print(f"  Row {ledger_row}: Made Vch Type bold: '{str(vch_type_cell.value)[:50]}...'")
                                 
                                 # Make all Debit and Credit values (Columns J and K) bold in this transaction block
                                 for bold_row in range(ledger_row, narration_row):
                                     # Make Debit column (J) bold
                                     debit_cell = worksheet.cell(row=bold_row, column=10)  # Column J (Debit)
                                     if debit_cell.value and debit_cell.value != '':
                                         debit_cell.font = bold_font
                                         print(f"  Row {bold_row}: Made Debit value bold: '{str(debit_cell.value)[:20]}...'")
                                     
                                     # Make Credit column (K) bold
                                     credit_cell = worksheet.cell(row=bold_row, column=11)  # Column K (Credit)
                                     if credit_cell.value and credit_cell.value != '':
                                         credit_cell.font = bold_font
                                         print(f"  Row {bold_row}: Made Credit value bold: '{str(credit_cell.value)[:20]}...'")
                                 
                                 break
            
            # Also check for "Opening Balance" text and make it bold along with its Debit/Credit values
            print("Checking for Opening Balance entries...")
            for row in range(10, worksheet.max_row + 1):
                cell_e = worksheet.cell(row=row, column=5)  # Column E (Description)
                
                # Check if this row contains "Opening Balance" text
                if (cell_e.value and 
                    isinstance(cell_e.value, str) and 
                    "Opening Balance" in str(cell_e.value)):
                    
                    # Make the Opening Balance text bold
                    cell_e.font = bold_font
                    print(f"  Row {row}: Made Opening Balance text bold: '{str(cell_e.value)[:50]}...'")
                    
                    # Make the associated Debit and Credit values bold
                    debit_cell = worksheet.cell(row=row, column=10)  # Column J (Debit)
                    if debit_cell.value and debit_cell.value != '':
                        debit_cell.font = bold_font
                        print(f"  Row {row}: Made Opening Balance Debit value bold: '{str(debit_cell.value)[:20]}...'")
                    
                    credit_cell = worksheet.cell(row=row, column=11)  # Column K (Credit)
                    if credit_cell.value and credit_cell.value != '':
                        credit_cell.font = bold_font
                        print(f"  Row {row}: Made Opening Balance Credit value bold: '{str(credit_cell.value)[:20]}...'")
            
            print("Output file transaction block formatting completed successfully!")
            
        except Exception as e:
            print(f"Error formatting output file transaction blocks: {e}")

    def create_matched_files(self, matches, transactions1, transactions2):
        """Create matched versions of both files with new columns."""
        if not matches:
            print("No matches found. Cannot create matched files.")
            return
        
        # Matches are already in sequential order from post-processing step
        print(f"\n=== USING SEQUENTIALLY ASSIGNED MATCHES ===")
        print(f"Total matches: {len(matches)}")
        print(f"First 10 Match IDs: {[m['match_id'] for m in matches[:10]]}")

        print(f"Last 10 Match IDs: {[m['match_id'] for m in matches[-10:]]}")
        
        # Create file1 with new columns
        file1_matched = transactions1.copy()
        
        # Create new columns with proper names
        match_id_col = pd.Series([None] * len(file1_matched), name='Match ID')
        audit_info_col = pd.Series([None] * len(file1_matched), name='Audit Info')
        match_type_col = pd.Series([None] * len(file1_matched), name='Match Type')
        
        # Concatenate new columns with existing data
        file1_matched = pd.concat([match_id_col, audit_info_col, file1_matched, match_type_col], axis=1)
        

        
        # Create file2 with new columns
        file2_matched = transactions2.copy()
        
        # Create new columns with proper names
        match_id_col2 = pd.Series([None] * len(file2_matched), name='Match ID')
        audit_info_col2 = pd.Series([None] * len(file2_matched), name='Audit Info')
        match_type_col2 = pd.Series([None] * len(file2_matched), name='Match Type')
        
        # Concatenate new columns with existing data
        file2_matched = pd.concat([match_id_col2, audit_info_col2, file2_matched, match_type_col2], axis=1)
        
        # print(f"DEBUG: File2 DataFrame created with shape: {file2_matched.shape}")
        # print(f"DEBUG: File2 columns: {list(file2_matched.columns)}")
        
        # print(f"DEBUG: Added Match Type column to both DataFrames")
        # print(f"DEBUG: File1 columns: {list(file1_matched.columns)}")
        # print(f"DEBUG: File2 columns: {list(file2_matched.columns)}")
        
        # Verify the new columns are actually there
        # print(f"DEBUG: File1 first few rows of Match ID column:")
        # print(file1_matched.iloc[:5, 0].tolist())
        # print(f"DEBUG: File1 first few rows of Audit Info column:")
        # print(file1_matched.iloc[:5, 1].tolist())
        
        print(f"\n=== DEBUG: MATCH DATA POPULATION ===")
        
        # Populate match information - process matches in sequential order
        for i, match in enumerate(matches):
            match_id = match['match_id']  # Use the pre-assigned match ID
            audit_info = self.create_audit_info(match)
            

            # Use the explicit Match_Type field if available, otherwise fall back to inference
            if 'Match_Type' in match and match['Match_Type']:
                match_type = match['Match_Type']
                print(f"  Match Type: {match_type} (from explicit field)")
            elif 'LC_Number' in match and match['LC_Number']:
                print(f"  LC Number: {match['LC_Number']}")
                match_type = 'LC'
            elif 'PO_Number' in match and match['PO_Number']:
                print(f"  PO Number: {match['PO_Number']}")
                match_type = 'PO'
            elif 'Interunit_Account' in match and match['Interunit_Account']:
                print(f"  Interunit Account: {match['Interunit_Account']}")
                match_type = 'Interunit'
            else:
                print(f"  Unknown Match Type")
                match_type = 'Unknown'
            print(f"  File1 Row {match['File1_Index']}: Debit={match['File1_Debit']}, Credit={match['File1_Credit']}")
            print(f"  File2 Row {match['File2_Index']}: Debit={match['File2_Debit']}, Credit={match['File2_Credit']}")
            print(f"  Audit Info: {audit_info}")
            print(f"  Match Type: {match_type}")
            
            # Update file1 - populate entire transaction block with Match ID and Audit Info
            file1_row_idx = match['File1_Index']
            print(f"    DEBUG: Setting File1 row {file1_row_idx} col 0 to '{match_id}'")
            print(f"    DEBUG: Setting File1 row {file1_row_idx} col 1 to '{audit_info[:50]}...'")
            print(f"    DEBUG: Setting File1 row {file1_row_idx} col -1 to '{match_type}' (last column)")
            
            # Find the entire transaction block for file1 and populate all rows
            file1_block_rows = self.block_identifier.get_transaction_block_rows(file1_row_idx, self.file1_path)
            print(f"    DEBUG: File1 transaction block spans rows: {file1_block_rows}")
            
            # Populate ALL rows of the transaction block with Match ID and Match Type, but only if not already set (preserve first/lowest Match ID)
            for i, block_row in enumerate(file1_block_rows):
                if 0 <= block_row < len(file1_matched):
                    # Only set Match ID if not already set (preserve first/lowest Match ID)
                    if pd.isna(file1_matched.iloc[block_row, 0]) or file1_matched.iloc[block_row, 0] == '':
                        file1_matched.iloc[block_row, 0] = match_id  # Match ID column (index 0)
                        file1_matched.iloc[block_row, -1] = match_type  # Match Type column (last column) - ALL ROWS
                    else:
                        print(f"      WARNING: Row {block_row} already has Match ID {file1_matched.iloc[block_row, 0]}, preserving it instead of {match_id}")
                    
                    # Audit Info goes ONLY in the second-to-last row of the transaction block
                    if i == len(file1_block_rows) - 2:  # Second-to-last row
                        file1_matched.iloc[block_row, 1] = audit_info  # Audit Info column (index 1)
                        print(f"    DEBUG: Populated File1 row {block_row} with Match ID '{match_id}', Audit Info, and Match Type '{match_type}' (second-to-last row)")
                    else:
                        print(f"    DEBUG: Populated File1 row {block_row} with Match ID '{match_id}' and Match Type '{match_type}'")
            

            
            # Update file2 - populate entire transaction block with Match ID and Audit Info
            file2_row_idx = match['File2_Index']
            print(f"    DEBUG: Setting File2 row {file2_row_idx} col 0 to '{match_id}'")
            print(f"    DEBUG: Setting File2 row {file2_row_idx} col 1 to '{audit_info[:50]}...'")
            print(f"    DEBUG: Setting File2 row {file2_row_idx} col -1 to '{match_type}' (last column)")
            
            # Find the entire transaction block for file2 and populate all rows
            file2_block_rows = self.block_identifier.get_transaction_block_rows(file2_row_idx, self.file2_path)
            print(f"    DEBUG: File2 transaction block spans rows: {file2_block_rows}")
            
            # Populate ALL rows of the transaction block with Match ID and Match Type, but only if not already set (preserve first/lowest Match ID)
            for i, block_row in enumerate(file2_block_rows):
                if 0 <= block_row < len(file2_matched):
                    # Only set Match ID if not already set (preserve first/lowest Match ID)
                    if pd.isna(file2_matched.iloc[block_row, 0]) or file2_matched.iloc[block_row, 0] == '':
                        file2_matched.iloc[block_row, 0] = match_id  # Match ID column (index 0)
                        file2_matched.iloc[block_row, -1] = match_type  # Match Type column (last column) - ALL ROWS
                    else:
                        print(f"      WARNING: Row {block_row} already has Match ID {file2_matched.iloc[block_row, 0]}, preserving it instead of {match_id}")
                    
                    # Audit Info goes ONLY in the second-to-last row of the transaction block
                    if i == len(file2_block_rows) - 2:  # Second-to-last row
                        file2_matched.iloc[block_row, 1] = audit_info  # Audit Info column (index 1)
                        print(f"    DEBUG: Populated File2 row {block_row} with Match ID '{match_id}', Audit Info, and Match Type '{match_type}' (second-to-last row)")
                    else:
                        print(f"    DEBUG: Populated File2 row {block_row} with Match ID '{match_id}' and Match Type '{match_type}'")
        
        # Save matched files using configuration variables
        base_name1 = os.path.splitext(os.path.basename(self.file1_path))[0]
        base_name2 = os.path.splitext(os.path.basename(self.file2_path))[0]
        
        # Get the directory of the input files
        input_dir1 = os.path.dirname(self.file1_path)
        input_dir2 = os.path.dirname(self.file2_path)
        
        output_file1 = os.path.join(input_dir1, f"{base_name1}{OUTPUT_SUFFIX}")
        output_file2 = os.path.join(input_dir2, f"{base_name2}{OUTPUT_SUFFIX}")
        
        print(f"\n=== OUTPUT FILE LOCATIONS ===")
        print(f"Input File 1: {self.file1_path}")
        print(f"Input Directory 1: {input_dir1}")
        print(f"Output File 1: {output_file1}")
        print(f"Input File 2: {self.file2_path}")
        print(f"Input Directory 2: {input_dir2}")
        print(f"Output File 2: {output_file2}")
        
        if VERBOSE_DEBUG:
            print(f"\n=== DEBUG: BEFORE SAVING ===")
            print(f"File1 - Rows with Match IDs: {file1_matched.iloc[:, 0].notna().sum()}")
            print(f"File1 - Rows with Audit Info: {file1_matched.iloc[:, 1].notna().sum()}")
            print(f"File1 - Rows with Match Type: {file1_matched.iloc[:, -1].notna().sum()}")
            print(f"File2 - Rows with Match IDs: {file2_matched.iloc[:, 0].notna().sum()}")
            print(f"File2 - Rows with Audit Info: {file2_matched.iloc[:, 1].notna().sum()}")
            print(f"File2 - Rows with Match Type: {file2_matched.iloc[:, -1].notna().sum()}")
            
            # Show some actual values to verify they're there
            print(f"\n=== DEBUG: ACTUAL VALUES IN DATAFRAME ===")
            
            # Get the actual populated rows dynamically
            populated_rows = file1_matched.iloc[:, 0].notna()
            if populated_rows.any():
                populated_indices = file1_matched[populated_rows].index
                for idx in populated_indices[:4]:  # Show first 4 populated rows
                    print(f"File1 - Row {idx} Match ID: '{file1_matched.iloc[idx, 0]}'")
                    print(f"File1 - Row {idx} Audit Info: '{file1_matched.iloc[idx, 1]}'")
            else:
                print("No populated rows found in File1")
        
        # Dates are already in string format - no conversion needed
        print("\n=== DATES ALREADY IN STRING FORMAT - NO CONVERSION NEEDED ===")
        
        # Create output with metadata + matched transactions
        with pd.ExcelWriter(output_file1, engine='openpyxl') as writer:
            # Write metadata
            self.metadata1.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            # Write matched transactions (skip first 8 rows to make room for metadata)
            file1_matched.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=8)
            
            # Get the worksheet to set column widths
            worksheet = writer.sheets['Sheet1']
            self._set_column_widths(worksheet)
            self._format_amount_columns(worksheet) # Apply amount formatting
            
            # Apply top alignment and text wrapping to ALL cells in the worksheet
            self._apply_top_alignment(worksheet)
            
            # Apply filters to the header row for easy data filtering and sorting
            self._apply_filters_to_header(worksheet)
            
            # Apply alternating background colors to matched transaction blocks
            self._apply_alternating_background_colors(worksheet, file1_matched)
            
            # Format output file transaction blocks (make narration italic)
            self._format_output_file_transaction_blocks(worksheet)
            
                    
        with pd.ExcelWriter(output_file2, engine='openpyxl') as writer:
            # Write metadata
            self.metadata2.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            # Write matched transactions (skip first 8 rows to make room for metadata)
            file2_matched.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=8)
            
            # Get the worksheet to set column widths
            worksheet = writer.sheets['Sheet1']
            self._set_column_widths(worksheet)
            self._format_amount_columns(worksheet) # Apply amount formatting
            
            # Apply top alignment and text wrapping to ALL cells in the worksheet
            self._apply_top_alignment(worksheet)
            
            # Apply filters to the header row for easy data filtering and sorting
            self._apply_filters_to_header(worksheet)
            
            # Apply alternating background colors to matched transaction blocks
            self._apply_alternating_background_colors(worksheet, file2_matched)
            
            # Format output file transaction blocks (make narration italic)
            self._format_output_file_transaction_blocks(worksheet)
            
        
        # Also create a simple version without metadata to test (if enabled)
        if CREATE_SIMPLE_FILES:
            simple_output1 = os.path.join(input_dir1, f"{base_name1}{SIMPLE_SUFFIX}")
            simple_output2 = os.path.join(input_dir2, f"{base_name2}{SIMPLE_SUFFIX}")
            
            print(f"\nCreating simple test files without metadata...")
            file1_matched.to_excel(simple_output1, index=False, header=True)
            file2_matched.to_excel(simple_output2, index=False, header=True)
            
            print(f"Created simple test files:")
            print(f"  {simple_output1}")
            print(f"  {simple_output2}")
        

        
        print(f"\n=== DEBUG: AFTER SAVING ===")
        print(f"Checking if files were actually written...")
        
        # Verify the files were written correctly
        try:
            df_check1 = pd.read_excel(output_file1, header=8)
            print(f"File1 loaded successfully, shape: {df_check1.shape}")
            print(f"File1 - Rows with Match IDs: {df_check1.iloc[:, 0].notna().sum()}")
            print(f"File1 - Rows with Audit Info: {df_check1.iloc[:, 1].notna().sum()}")
            print(f"File1 - Rows with Match Type: {df_check1.iloc[:, -1].notna().sum()}")
            
            # Check if text wrapping was applied by reading the Excel file with openpyxl
            print(f"\n=== VERIFYING TEXT WRAPPING IN FILE 1 ===")
            wb1 = openpyxl.load_workbook(output_file1)
            ws1 = wb1.active
            print(f"Worksheet: {ws1.title}")
            print(f"Max row: {ws1.max_row}, Max column: {ws1.max_column}")
            
            # Check a few cells in columns B and E for text wrapping
            for row in range(9, min(15, ws1.max_row + 1)):
                cell_b = ws1.cell(row=row, column=2)
                cell_e = ws1.cell(row=row, column=5)
                print(f"Row {row}:")
                print(f"  Column B: value='{cell_b.value}', wrap_text={cell_b.alignment.wrap_text if cell_b.alignment else 'None'}")
                print(f"  Column E: value='{cell_e.value}', wrap_text={cell_e.alignment.wrap_text if cell_e.alignment else 'None'}")
            
            wb1.close()
            
        except Exception as e:
            print(f"Error reading File1: {e}")
        
        try:
            df_check2 = pd.read_excel(output_file2, header=8)
            print(f"File2 loaded successfully, shape: {df_check2.shape}")
            print(f"File2 - Rows with Match IDs: {df_check2.iloc[:, 0].notna().sum()}")
            print(f"File2 - Rows with Audit Info: {df_check2.iloc[:, 1].notna().sum()}")
            print(f"File2 - Rows with Match Type: {df_check2.iloc[:, -1].notna().sum()}")
            
            # Check if text wrapping was applied by reading the Excel file with openpyxl
            print(f"\n=== VERIFYING TEXT WRAPPING IN FILE 2 ===")
            wb2 = openpyxl.load_workbook(output_file2)
            ws2 = wb2.active
            print(f"Worksheet: {ws2.title}")
            print(f"Max row: {ws2.max_row}, Max column: {ws2.max_column}")
            
            # Check a few cells in columns B and E for text wrapping
            for row in range(9, min(15, ws2.max_row + 1)):
                cell_b = ws2.cell(row=row, column=2)
                cell_e = ws2.cell(row=row, column=5)
                print(f"Row {row}:")
                print(f"  Column B: value='{cell_b.value}', wrap_text={cell_b.alignment.wrap_text if cell_b.alignment else 'None'}")
                print(f"  Column E: value='{cell_e.value}', wrap_text={cell_e.alignment.wrap_text if cell_e.alignment else 'None'}")
            
            wb2.close()
            
        except Exception as e:
            print(f"Error reading File2: {e}")
        
        print(f"\nCreated matched files:")
        print(f"  {output_file1}")
        print(f"  {output_file2}")
        
        # Verify the data was populated correctly
        self.verify_match_data(file1_matched, file2_matched)
    

    
    def verify_match_data(self, file1_matched, file2_matched):
        """Verify that Match ID and Audit Info columns are properly populated."""
        print(f"\n=== VERIFICATION RESULTS ===")
        
        # Check Match ID column population - only count non-empty, non-NaN values
        match_ids_1 = file1_matched.iloc[:, 0].replace('', None).dropna()
        match_ids_2 = file2_matched.iloc[:, 0].replace('', None).dropna()
        
        print(f"File 1 - Match IDs populated: {len(match_ids_1)}")
        print(f"File 2 - Match IDs populated: {len(match_ids_2)}")
        
        # Check Audit Info column population - only count non-empty, non-NaN values
        audit_info_1 = file1_matched.iloc[:, 1].replace('', None).dropna()
        audit_info_2 = file2_matched.iloc[:, 1].replace('', None).dropna()
        
        print(f"File 1 - Audit Info populated: {len(audit_info_1)}")
        print(f"File 2 - Audit Info populated: {len(audit_info_2)}")
        
        # Check Match Type column population - only count non-empty, non-NaN values
        match_types_1 = file1_matched.iloc[:, -1].replace('', None).dropna()
        match_types_2 = file2_matched.iloc[:, -1].replace('', None).dropna()
        
        print(f"File 1 - Match Types populated: {len(match_types_1)}")
        print(f"File 2 - Match Types populated: {len(match_types_2)}")
        
        # Show sample populated data
        if len(match_ids_1) > 0:
            print(f"\nFile 1 - Sample populated rows:")
            for _, match_id in enumerate(match_ids_1[:3]):
                # Find rows that actually have this match_id (not empty strings)
                mask = (file1_matched.iloc[:, 0] == match_id) & (file1_matched.iloc[:, 0] != '')
                if mask.any():
                    row_idx = file1_matched[mask].index[0]
                    print(f"  Row {row_idx}: Match ID = {match_id}")
                    print(f"    Date: {file1_matched.iloc[row_idx, 3]}")
                    print(f"    Description: {str(file1_matched.iloc[row_idx, 4])[:50]}...")
                    print(f"    Debit: {file1_matched.iloc[row_idx, 10]}, Credit: {file1_matched.iloc[row_idx, 11]}")
                    print(f"    Match Type: {file1_matched.iloc[row_idx, -1]}")
        
        if len(match_ids_2) > 0:
            print(f"\nFile 2 - Sample populated rows:")
            for i, match_id in enumerate(match_ids_2[:3]):
                # Find rows that actually have this match_id (not empty strings)
                mask = (file2_matched.iloc[:, 0] == match_id) & (file2_matched.iloc[:, 0] != '')
                if mask.any():
                    row_idx = file2_matched[mask].index[0]
                    print(f"  Row {row_idx}: Match ID = {match_id}")
                    print(f"    Date: {file2_matched.iloc[row_idx, 3]}")
                    print(f"    Description: {str(file2_matched.iloc[row_idx, 4])[:50]}...")
                    print(f"    Debit: {file2_matched.iloc[row_idx, 10]}, Credit: {file2_matched.iloc[row_idx, 11]}")
                    print(f"    Match Type: {file2_matched.iloc[row_idx, -1]}")

def main():
    # Show current configuration
    print_configuration()
    print()
    
    # Use configuration variables from the top of the file
    print(f"=== PROCESSING FILES ===")
    
    # Check if input files exist
    if not os.path.exists(INPUT_FILE1_PATH):
        print(f"ERROR: Input file 1 not found: {INPUT_FILE1_PATH}")
        return
    if not os.path.exists(INPUT_FILE2_PATH):
        print(f"ERROR: Input file 2 not found: {INPUT_FILE2_PATH}")
        return
    
    # Create output folder if it doesn't exist
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    # Create matcher instance
    matcher = ExcelTransactionMatcher(INPUT_FILE1_PATH, INPUT_FILE2_PATH)
    matches = matcher.find_potential_matches()
    # Get the data from the matcher instance
    transactions1 = matcher.transactions1
    transactions2 = matcher.transactions2
    
    print(f"\n=== SUMMARY ===")
    print(f"Total potential matches found: {len(matches)}")
    
    if matches:
        print("\nCreating matched output files...")
        matcher.create_matched_files(matches, transactions1, transactions2)
        print("\nOutput files created successfully!")
    else:
        print("\nNo matches found. No output files created.")

if __name__ == "__main__":
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Excel Transaction Matcher for LC Numbers')
    parser.add_argument('--file1', help='Path to first Excel file (overrides config)')
    parser.add_argument('--file2', help='Path to second Excel file (overrides config)')
    parser.add_argument('--output', help='Output folder (overrides config)')
    parser.add_argument('--config', action='store_true', help='Show current configuration')
    parser.add_argument('--help-config', action='store_true', help='Show configuration help')
    
    args = parser.parse_args()
    
    # Handle special arguments
    if args.help_config:
        update_configuration()
        sys.exit(0)
    
    if args.config:
        print_configuration()
        sys.exit(0)
    
    # Override configuration with command line arguments if provided
    if args.file1:
        INPUT_FILE1_PATH = args.file1
    if args.file2:
        INPUT_FILE2_PATH = args.file2
    if args.output:
        OUTPUT_FOLDER = args.output
    
    # Run the main function
    main()
    